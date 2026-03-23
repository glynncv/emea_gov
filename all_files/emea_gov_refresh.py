"""
EMEA Governance Cockpit 2026 — SNOW REST API Refresh Script
Version: v1
Owner: EMEA SDM (Colman)
Purpose: Query SNOW REST API, calculate Operations Panel metrics,
         update Physics Engine trends, write results to cockpit Excel file.

Dependencies:
    pip install requests openpyxl pandas python-dotenv

Usage:
    python emea_gov_refresh.py

Configuration:
    Create a .env file in the same directory (see CONFIG section below).
    Never commit credentials to source control.
"""

import os
import requests
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime, timedelta, timezone
from dotenv import load_dotenv

load_dotenv()

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
# .env file should contain:
# SNOW_INSTANCE=your-instance.service-now.com
# SNOW_USER=your_username
# SNOW_PASS=your_password
# COCKPIT_PATH=C:\EMEA_GOV\Cockpit\EMEA_Governance_Cockpit_2026.xlsx

SNOW_INSTANCE = os.getenv("SNOW_INSTANCE")
SNOW_USER     = os.getenv("SNOW_USER")
SNOW_PASS     = os.getenv("SNOW_PASS")
COCKPIT_PATH  = os.getenv("COCKPIT_PATH", "EMEA_Governance_Cockpit_2026.xlsx")

BASE_URL = f"https://{SNOW_INSTANCE}/api/now/table"
AUTH     = (SNOW_USER, SNOW_PASS)
HEADERS  = {"Accept": "application/json", "Content-Type": "application/json"}

TODAY = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)

# SLA targets in hours — used if sla_target field absent from SNOW
SLA_DEFAULTS = {"1": 4, "2": 8, "3": 72, "4": 336}

# 25 EMEA physical sites — exact SNOW u_site_name values
EMEA_SITES = [
    "Duesseldorf - Germany",
    "Warwick (Titan) - United Kingdom",
    "Madrid - Spain",
    "Blois - France",
    "Cinisello - Italy",
    "Izmir - Turkey / ESBAS 3 (PT Phase 2)",
    "Izmir - Turkey / ESBAS 2 (PT Phase 1)",
    "Gillingham - United Kingdom",
    "Stonehouse - United Kingdom",
    "Iasi - Romania",
    "Bucharest - Romania",
    "Rzeszow - Poland",
    "Dubai - United Arab Emirates",
    "Amal - Sweden",
    "Istanbul - Turkey",
    "Cergy - France",
    "Warsaw (Delphi Academy) - Poland",
    "Warsaw - Poland",
    "Technical Center Krakow",
    "Krakow - Poland",
    "Hartridge - United Kingdom",
    "Belval - Luxembourg",
    "Wroclaw - Poland",
    "Blonie - Poland",
    "Warwick - United Kingdom",
]

SITE_FILTER = "^u_site_nameIN" + ",".join(EMEA_SITES)


# ─────────────────────────────────────────────
# SNOW API HELPERS
# ─────────────────────────────────────────────

def snow_query(table: str, query: str, fields: list[str],
               limit: int = 10000) -> pd.DataFrame:
    """
    Query a SNOW table via REST API. Returns a DataFrame.
    Handles pagination automatically if record count exceeds limit.
    """
    params = {
        "sysparm_query":  query,
        "sysparm_fields": ",".join(fields),
        "sysparm_limit":  limit,
        "sysparm_offset": 0,
    }
    records = []
    while True:
        resp = requests.get(
            f"{BASE_URL}/{table}",
            auth=AUTH, headers=HEADERS, params=params
        )
        resp.raise_for_status()
        batch = resp.json().get("result", [])
        records.extend(batch)
        if len(batch) < limit:
            break
        params["sysparm_offset"] += limit

    df = pd.DataFrame(records)
    if df.empty:
        print(f"  WARNING: No records returned for {table} — check site filter")
        return df

    # Flatten dot-walked fields e.g. {"value": "...", "display_value": "..."}
    for col in df.columns:
        if df[col].dtype == object:
            sample = df[col].dropna().iloc[0] if not df[col].dropna().empty else None
            if isinstance(sample, dict) and "value" in sample:
                df[col] = df[col].apply(
                    lambda x: x.get("display_value") or x.get("value")
                    if isinstance(x, dict) else x
                )

    # Parse date fields
    for col in ["opened_at", "sys_updated_on", "resolved_at", "closed_at"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], utc=True, errors="coerce")

    return df


# ─────────────────────────────────────────────
# DATA RETRIEVAL
# ─────────────────────────────────────────────
# EUC / EOSL — reads from PBI Excel export, NOT SNOW
# ─────────────────────────────────────────────

def fetch_euc_assets(euc_path: str) -> pd.DataFrame:
    """
    Read EOSL device data from EUC_EOSL.xlsx (Power BI export — not SNOW).
    Reads the 'ESOL Replacement Tracker' sheet site matrix.
    Falls back to the raw 'Export' sheet if Tracker sheet is absent.

    Args:
        euc_path: Full path to EUC_EOSL.xlsx

    Returns:
        DataFrame with columns: site, target_date_label, device_count
        Summary dict: {total: int, overdue: int, by_nov_2026: int}
    """
    from openpyxl import load_workbook as _lw
    try:
        wb = _lw(euc_path, read_only=True, data_only=True)

        if "ESOL Replacement Tracker" in wb.sheetnames:
            ws = wb["ESOL Replacement Tracker"]
            rows = list(ws.iter_rows(values_only=True))
            data, header = [], None
            for row in rows:
                if row[0] and "site" in str(row[0]).lower():
                    header = row
                    continue
                if header and row[0]:
                    site = str(row[0]).strip()
                    for i, val in enumerate(row[1:], 1):
                        if val and header[i]:
                            data.append({
                                "site": site,
                                "target_date_label": str(header[i]),
                                "device_count": int(val) if val else 0
                            })
            if data:
                print(f"  EUC: {len(data)} site/target rows from ESOL Replacement Tracker.")
                return pd.DataFrame(data)

        # Fallback: raw Export sheet
        if "Export" in wb.sheetnames:
            ws = wb["Export"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h).strip() if h else "" for h in rows[0]]
                df = pd.DataFrame(
                    [dict(zip(headers, r)) for r in rows[1:] if any(r)]
                )
                site_col = next((c for c in df.columns if "site" in c.lower()), None)
                if site_col:
                    counts = df[site_col].value_counts().reset_index()
                    counts.columns = ["site", "device_count"]
                    counts["target_date_label"] = "All EOSL"
                    print(f"  EUC: Fallback — {len(df)} raw device records from Export sheet.")
                    return counts

        print("  EUC: WARNING — could not read EUC_EOSL.xlsx. Check sheet names.")
        return pd.DataFrame()

    except FileNotFoundError:
        print(f"  EUC: WARNING — file not found: {euc_path}. Physics Block 1 skipped.")
        return pd.DataFrame()
    except Exception as e:
        print(f"  EUC: ERROR reading {euc_path}: {e}")
        return pd.DataFrame()


# ─────────────────────────────────────────────

def fetch_incidents() -> pd.DataFrame:
    """Open incidents for all 25 EMEA sites."""
    query  = SITE_FILTER + "^stateNOT IN6,7"  # 6=Resolved, 7=Closed
    fields = [
        "number", "u_site_name", "priority",
        "opened_at", "sys_updated_on",
        "sla_target", "cmdb_ci", "problem_id",
        "state", "short_description"
    ]
    print("  Fetching open incidents...")
    return snow_query("incident", query, fields)


def fetch_major_incident_history() -> pd.DataFrame:
    """P1/P2 incidents (open or closed) in the past 60 days — for repeat MI check."""
    since = (TODAY - timedelta(days=60)).strftime("%Y-%m-%d %H:%M:%S")
    query = (SITE_FILTER
             + "^priorityIN1,2"
             + f"^opened_at>={since}")
    fields = [
        "number", "u_site_name", "priority",
        "opened_at", "resolved_at", "closed_at",
        "cmdb_ci", "problem_id", "state", "short_description"
    ]
    print("  Fetching MI history (60 days)...")
    return snow_query("incident", query, fields)


def fetch_catalogue_tasks() -> pd.DataFrame:
    """Open catalogue tasks for all 25 EMEA sites."""
    query  = SITE_FILTER + "^stateNOT IN4,7"  # 4=Closed Complete, 7=Cancelled
    fields = [
        "number", "u_site_name",
        "opened_at", "sys_updated_on",
        "state", "short_description"
    ]
    print("  Fetching open catalogue tasks...")
    return snow_query("sc_task", query, fields)


def fetch_problems() -> pd.DataFrame:
    """Open problems without RCA for all 25 EMEA sites."""
    query  = SITE_FILTER + "^stateNOT IN4,7"
    fields = [
        "number", "u_site_name",
        "opened_at", "sys_updated_on",
        "u_root_cause", "root_cause",
        "state", "short_description", "assigned_to"
    ]
    print("  Fetching open problems...")
    df = snow_query("problem", query, fields)

    # Normalise root cause field — instance may use either name
    if "u_root_cause" in df.columns and "root_cause" not in df.columns:
        df = df.rename(columns={"u_root_cause": "root_cause"})
    elif "u_root_cause" in df.columns:
        df["root_cause"] = df["root_cause"].fillna(df["u_root_cause"])

    return df


# ─────────────────────────────────────────────
# METRIC CALCULATIONS
# ─────────────────────────────────────────────

def age_days(df: pd.DataFrame, date_col: str = "opened_at") -> pd.Series:
    """Return age in days from date_col to today."""
    return (TODAY - df[date_col]).dt.total_seconds() / 86400


def calc_m1_incident_aging(incidents: pd.DataFrame) -> dict:
    """Metric 1: % open incidents aged <= 10 days."""
    if incidents.empty:
        return {"value": None, "note": "No data"}
    incidents = incidents.copy()
    incidents["age_days"] = age_days(incidents)
    compliant = (incidents["age_days"] <= 10).sum()
    total     = len(incidents)
    pct       = round(compliant / total * 100, 1) if total else 0
    return {
        "value": pct,
        "note":  f"{compliant}/{total} incidents aged <=10 days"
    }


def calc_m2_catalogue_aging(tasks: pd.DataFrame) -> dict:
    """Metric 2: % open catalogue tasks aged <= 30 days."""
    if tasks.empty:
        return {"value": None, "note": "No data"}
    tasks = tasks.copy()
    tasks["age_days"] = age_days(tasks)
    compliant = (tasks["age_days"] <= 30).sum()
    total     = len(tasks)
    pct       = round(compliant / total * 100, 1) if total else 0
    return {
        "value": pct,
        "note":  f"{compliant}/{total} tasks aged <=30 days"
    }


def calc_m3_sla_x2(incidents: pd.DataFrame) -> dict:
    """Metric 3: Count of open incidents exceeding SLA x2, split by priority."""
    if incidents.empty:
        return {"value": 0, "note": "No data", "p12_count": 0, "p34_count": 0}
    inc = incidents.copy()
    inc["age_days"] = age_days(inc)

    # SLA target in days per record
    if "sla_target" in inc.columns and inc["sla_target"].notna().any():
        inc["sla_days"] = pd.to_numeric(inc["sla_target"], errors="coerce") / 24
    else:
        inc["sla_days"] = inc["priority"].map(
            {k: v / 24 for k, v in SLA_DEFAULTS.items()}
        )

    inc["sla_x2"] = inc["age_days"] > (inc["sla_days"] * 2)
    breached = inc[inc["sla_x2"]]
    p12 = breached[breached["priority"].isin(["1", "2"])].shape[0]
    p34 = breached[breached["priority"].isin(["3", "4"])].shape[0]
    flag = " ⚠ P1/P2 PRESENT" if p12 > 0 else ""
    return {
        "value":    p12 + p34,
        "note":     f"P1/P2: {p12} | P3/P4: {p34}{flag}",
        "p12_count": p12,
        "p34_count": p34,
    }


def calc_m4_no_movement(incidents: pd.DataFrame,
                        tasks: pd.DataFrame) -> dict:
    """Metric 4: Count of tickets with no update for >= 14 days."""
    results = []
    site_counts = {}

    for df, label in [(incidents, "INC"), (tasks, "CAT")]:
        if df.empty:
            continue
        d = df.copy()
        d["stale_days"] = (TODAY - d["sys_updated_on"]).dt.total_seconds() / 86400
        stale = d[d["stale_days"] >= 14]
        results.append(len(stale))
        for site, cnt in stale["u_site_name"].value_counts().items():
            site_counts[site] = site_counts.get(site, 0) + cnt

    total = sum(results)
    top5  = sorted(site_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    top5_str = ", ".join(f"{s}:{c}" for s, c in top5) if top5 else "none"
    return {
        "value": total,
        "note":  f"Top sites: {top5_str}"
    }


def calc_m5_repeat_mi(mi_history: pd.DataFrame) -> dict:
    """Metric 5: Repeat MIs for same CI or site within any 30-day window."""
    if mi_history.empty:
        return {"value": 0, "note": "No data", "watch_count": 0, "breach_count": 0}

    mi = mi_history.copy().sort_values("opened_at")
    watch_items   = []
    breach_items  = []

    for (site, ci), grp in mi.groupby(["u_site_name", "cmdb_ci"]):
        dates = grp["opened_at"].tolist()
        for i in range(len(dates)):
            for j in range(i + 1, len(dates)):
                delta = abs((dates[j] - dates[i]).days)
                if delta <= 30:
                    no_problem = grp["problem_id"].isna().any() or \
                                 (grp["problem_id"] == "").any()
                    entry = f"{site} / CI: {ci or 'unknown'}"
                    if no_problem:
                        watch_items.append(entry)
                    else:
                        breach_items.append(entry)

    watch_items  = list(set(watch_items))
    breach_items = list(set(breach_items))

    note_parts = []
    if breach_items:
        note_parts.append("BREACH: " + "; ".join(breach_items[:3]))
    if watch_items:
        note_parts.append("WATCH: " + "; ".join(watch_items[:3]))

    return {
        "value":        len(watch_items) + len(breach_items),
        "note":         " | ".join(note_parts) if note_parts else "No repeats",
        "watch_count":  len(watch_items),
        "breach_count": len(breach_items),
    }


def calc_m6_problems_no_rca(problems: pd.DataFrame) -> dict:
    """Metric 6: Open problems with no RCA and age > 30 days."""
    if problems.empty:
        return {"value": 0, "note": "No data", "watch_count": 0, "breach_count": 0}

    p = problems.copy()
    p["age_days"] = age_days(p)

    no_rca = p[
        p["age_days"] > 30
    ].copy()

    if "root_cause" in no_rca.columns:
        no_rca = no_rca[
            no_rca["root_cause"].isna() | (no_rca["root_cause"] == "")
        ]

    watch  = no_rca[(no_rca["age_days"] > 30) & (no_rca["age_days"] <= 60)]
    breach = no_rca[no_rca["age_days"] > 60]

    return {
        "value":        len(no_rca),
        "note":         f"Watch (30-60d): {len(watch)} | Breach (60d+): {len(breach)}",
        "watch_count":  len(watch),
        "breach_count": len(breach),
    }


# ─────────────────────────────────────────────
# TRIGGER STATUS EVALUATION
# ─────────────────────────────────────────────

def evaluate_trigger(metric_name: str, current: float,
                     prev_4w_avg: float, prev_week: float,
                     extra: dict = None) -> str:
    """
    Returns GREEN / WATCH / BREACHED based on trigger rules.
    extra: dict of additional context (p12_count, watch_count, breach_count etc.)
    """
    extra = extra or {}

    if metric_name == "incident_aging":
        if current is not None and current < 90 and prev_4w_avg is not None and prev_4w_avg < 90:
            return "BREACHED"
        if current is not None and current < 90:
            return "WATCH"
        return "GREEN"

    elif metric_name == "catalogue_aging":
        if current is not None and current < 90 and prev_week is not None and prev_week < 90:
            return "BREACHED"
        if current is not None and current < 90:
            return "WATCH"
        return "GREEN"

    elif metric_name == "no_movement":
        if current and current > 0 and prev_week and current > prev_week:
            return "BREACHED"  # increasing for >=2 weeks — simplification: upward trend
        if current and current > 0:
            return "WATCH"
        return "GREEN"

    elif metric_name == "repeat_mi":
        if extra.get("breach_count", 0) > 0:
            return "BREACHED"
        if extra.get("watch_count", 0) > 0:
            return "WATCH"
        return "GREEN"

    elif metric_name == "problems_no_rca":
        if extra.get("breach_count", 0) > 0:
            return "BREACHED"
        if extra.get("watch_count", 0) > 0:
            return "WATCH"
        return "GREEN"

    elif metric_name == "sla_x2":
        if extra.get("p12_count", 0) > 0:
            return "BREACHED"
        if current and current > 0 and prev_week and current > prev_week:
            return "BREACHED"
        if current and current > 0:
            return "WATCH"
        return "GREEN"

    return "PENDING"


# ─────────────────────────────────────────────
# EXCEL WRITE
# ─────────────────────────────────────────────

# Trigger status colour fills
FILLS = {
    "GREEN":   PatternFill("solid", fgColor="00B050"),
    "WATCH":   PatternFill("solid", fgColor="FFC000"),
    "BREACHED":PatternFill("solid", fgColor="FF0000"),
    "PENDING": PatternFill("solid", fgColor="D9D9D9"),
}
FONTS = {
    "GREEN":    Font(bold=True, color="FFFFFF"),
    "WATCH":    Font(bold=True, color="000000"),
    "BREACHED": Font(bold=True, color="FFFFFF"),
    "PENDING":  Font(bold=False, color="000000"),
}


def set_trigger_cell(ws, cell_ref: str, status: str):
    cell = ws[cell_ref]
    cell.value = status
    cell.fill  = FILLS.get(status, FILLS["PENDING"])
    cell.font  = FONTS.get(status, FONTS["PENDING"])


def shift_physics_trends(ws_physics, row: int, new_value: float):
    """
    Shift Wk1→Wk2, Wk2→Wk3, Wk3→Wk4 in Physics_Engine Block 4.
    Then write new_value into Wk1.
    Column layout assumed: A=Metric, B=Wk1, C=Wk2, D=Wk3, E=Wk4
    (adjust col letters to match your actual workbook layout)
    """
    cols = ["B", "C", "D", "E"]  # Wk1, Wk2, Wk3, Wk4
    # Shift right to left: E=D, D=C, C=B
    for i in range(len(cols) - 1, 0, -1):
        src = ws_physics[f"{cols[i-1]}{row}"].value
        ws_physics[f"{cols[i]}{row}"] = src
    ws_physics[f"{cols[0]}{row}"] = new_value  # new value into Wk1


def update_cockpit(metrics: dict, prev_values: dict):
    """
    Write all calculated metrics into the cockpit Excel file.
    Row references below match the Operations_Panel layout from Build Pack v1.
    Adjust row numbers if your build came out differently.
    """
    wb = openpyxl.load_workbook(COCKPIT_PATH)
    ws_ops     = wb["Operations_Panel"]
    ws_physics = wb["Physics_Engine"]

    # ── Row mapping: metric → (Current Value col, Trigger Status col,
    #                           Escalation Required col, Physics Block 4 row)
    # Adjust row numbers to match your actual workbook after build
    ROW_MAP = {
        "incident_aging":   {"cv": "D3",  "ts": "H3",  "er": "I3",  "phys_row": 2},
        "catalogue_aging":  {"cv": "D4",  "ts": "H4",  "er": "I4",  "phys_row": 3},
        "sla_x2":           {"cv": "D5",  "ts": "H5",  "er": "I5",  "phys_row": 4},
        "no_movement":      {"cv": "D7",  "ts": "H7",  "er": "I7",  "phys_row": 5},
        "repeat_mi":        {"cv": "D9",  "ts": "H9",  "er": "I9",  "phys_row": 6},
        "problems_no_rca":  {"cv": "D10", "ts": "H10", "er": "I10", "phys_row": 7},
    }

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    results_summary = []

    for key, m in metrics.items():
        if key not in ROW_MAP:
            continue
        refs = ROW_MAP[key]

        # Current value
        cv_cell = ws_ops[refs["cv"]]
        cv_cell.value = m.get("value")

        # Cell note / comment (openpyxl Comment requires author)
        from openpyxl.comments import Comment
        if m.get("note"):
            cv_cell.comment = Comment(
                f"{now_str}: {m['note']}", "EMEA SDM Refresh"
            )

        # Trigger status
        prev_cv   = prev_values.get(key, {}).get("current")
        prev_4w   = prev_values.get(key, {}).get("avg_4w")
        status    = evaluate_trigger(
            key,
            current    = m.get("value"),
            prev_4w_avg= prev_4w,
            prev_week  = prev_cv,
            extra      = m
        )
        set_trigger_cell(ws_ops, refs["ts"], status)

        # Escalation Required
        esc_cell = ws_ops[refs["er"]]
        esc_cell.value = "YES" if status in ("WATCH", "BREACHED") else "NO"
        esc_cell.fill  = FILLS["BREACHED"] if status in ("WATCH", "BREACHED") \
                         else FILLS["GREEN"]
        esc_cell.font  = FONTS["BREACHED"] if status in ("WATCH", "BREACHED") \
                         else FONTS["GREEN"]

        # Physics Engine Block 4 trend shift
        shift_physics_trends(
            ws_physics, refs["phys_row"], m.get("value")
        )

        results_summary.append({
            "metric": key,
            "value":  m.get("value"),
            "status": status,
            "note":   m.get("note", "")
        })

    wb.save(COCKPIT_PATH)
    return results_summary


# ─────────────────────────────────────────────
# PREVIOUS VALUES (for trend evaluation)
# ─────────────────────────────────────────────

def read_prev_values() -> dict:
    """
    Read current cockpit values before overwriting — used for trend comparison.
    Returns dict of {metric_key: {current, avg_4w}}
    Adjust cell refs to match your workbook.
    """
    try:
        wb = openpyxl.load_workbook(COCKPIT_PATH, data_only=True)
        ws = wb["Operations_Panel"]
        return {
            "incident_aging":  {"current": ws["D3"].value,  "avg_4w": ws["E3"].value},
            "catalogue_aging": {"current": ws["D4"].value,  "avg_4w": ws["E4"].value},
            "sla_x2":          {"current": ws["D5"].value,  "avg_4w": ws["E5"].value},
            "no_movement":     {"current": ws["D7"].value,  "avg_4w": ws["E7"].value},
            "repeat_mi":       {"current": ws["D9"].value,  "avg_4w": ws["E9"].value},
            "problems_no_rca": {"current": ws["D10"].value, "avg_4w": ws["E10"].value},
        }
    except Exception:
        return {}


# ─────────────────────────────────────────────
# SITE COVERAGE VALIDATION
# ─────────────────────────────────────────────

def validate_site_coverage(incidents: pd.DataFrame,
                            tasks: pd.DataFrame,
                            problems: pd.DataFrame):
    """Warn if any of the 25 expected sites are missing from any dataset."""
    all_sites = set()
    for df in [incidents, tasks, problems]:
        if not df.empty and "u_site_name" in df.columns:
            all_sites.update(df["u_site_name"].dropna().unique())

    missing  = set(EMEA_SITES) - all_sites
    extra    = all_sites - set(EMEA_SITES)

    if missing:
        print(f"  ⚠  MISSING from data: {sorted(missing)}")
        print("     Check SNOW site filter — these sites returned no records")
    if extra:
        print(f"  ⚠  UNEXPECTED sites in data: {sorted(extra)}")
        print("     SNOW filter may be too broad")
    if not missing and not extra:
        print("  ✓  All 25 EMEA sites present in data")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    print(f"\n{'='*60}")
    print(f"EMEA Governance Cockpit — Weekly Refresh")
    print(f"Run: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'='*60}\n")

    # 1. Read previous cockpit values for trend comparison
    print("Reading previous cockpit values...")
    prev_values = read_prev_values()

    # 2. Fetch data from SNOW
    print("\nQuerying SNOW REST API...")
    incidents  = fetch_incidents()
    mi_history = fetch_major_incident_history()
    tasks      = fetch_catalogue_tasks()
    problems   = fetch_problems()

    print(f"\n  Incidents (open):    {len(incidents):>5} records")
    print(f"  MI history (60d):    {len(mi_history):>5} records")
    print(f"  Catalogue tasks:     {len(tasks):>5} records")
    print(f"  Problems:            {len(problems):>5} records")

    # 3. Validate site coverage
    print("\nValidating site coverage...")
    validate_site_coverage(incidents, tasks, problems)

    # 4. Calculate metrics
    print("\nCalculating metrics...")
    metrics = {
        "incident_aging":  calc_m1_incident_aging(incidents),
        "catalogue_aging": calc_m2_catalogue_aging(tasks),
        "sla_x2":          calc_m3_sla_x2(incidents),
        "no_movement":     calc_m4_no_movement(incidents, tasks),
        "repeat_mi":       calc_m5_repeat_mi(mi_history),
        "problems_no_rca": calc_m6_problems_no_rca(problems),
    }

    # 5. Update cockpit
    print(f"\nUpdating cockpit: {COCKPIT_PATH}")
    summary = update_cockpit(metrics, prev_values)

    # 6. Print summary
    print(f"\n{'─'*60}")
    print("REFRESH SUMMARY")
    print(f"{'─'*60}")
    breached = [r for r in summary if r["status"] == "BREACHED"]
    watch    = [r for r in summary if r["status"] == "WATCH"]

    for r in summary:
        icon = "🔴" if r["status"] == "BREACHED" \
               else "🟡" if r["status"] == "WATCH" \
               else "🟢"
        print(f"  {icon} {r['metric']:<22} {str(r['value']):<10} {r['status']}")
        if r["note"]:
            print(f"       {r['note']}")

    print(f"\n  BREACHED: {len(breached)} | WATCH: {len(watch)} "
          f"| GREEN: {len(summary) - len(breached) - len(watch)}")

    if breached:
        print("\n  ⚠  BREACHED metrics require immediate escalation:")
        for r in breached:
            print(f"     → {r['metric']}: {r['note']}")

    print(f"\n{'='*60}")
    print("Cockpit saved. Refresh complete.")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
