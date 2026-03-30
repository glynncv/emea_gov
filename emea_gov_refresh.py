"""
EMEA Governance Cockpit 2026 — SNOW REST API Refresh Script
Version: v5
Owner: EMEA SDM (Colman)
Purpose: Query SNOW REST API, calculate Operations Panel metrics,
         update Physics Engine trends, write results to cockpit Excel file.
         Open incidents and MI history use batch-by-site queries (location.u_site_name
         per EMEA site) because this SNOW instance ignores broader location filters on incident.

Dependencies:
    pip install requests openpyxl pandas python-dotenv

Usage:
    python emea_gov_refresh.py                          # normal API mode
    python emea_gov_refresh.py --dry-run                # Test run - no cockpit update
    python emea_gov_refresh.py --csv                    # CSV fallback, default path
    python emea_gov_refresh.py --csv --csv-path "C:/path/to/file.csv"  # explicit path
    python emea_gov_refresh.py --csv --dry-run          # CSV load, no cockpit write
    python emea_gov_refresh.py --test-fetch             # Test incident fetch strategies
    python emea_gov_refresh.py --shadow-backfill       # One-time historical Wk2/Wk3 Physics Block 4 backfill

Configuration:
    Create a .env file in the same directory (see CONFIG section below).
    Never commit credentials to source control.
"""

import os
import re
import sys
import time
import argparse
import logging
from logging.handlers import TimedRotatingFileHandler
import requests
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime, timedelta, timezone
from dotenv import load_dotenv

load_dotenv()

# ---------------------------------------------
# LOGGING SETUP
# ---------------------------------------------
LOG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
os.makedirs(LOG_DIR, exist_ok=True)
LOG_FILE = os.path.join(LOG_DIR, "refresh_log.txt")

logger = logging.getLogger("emea_gov_refresh")
logger.setLevel(logging.INFO)

# Weekly rotating log handler (keeps 8 weeks of history)
file_handler = TimedRotatingFileHandler(
    LOG_FILE,
    when="W0",  # Rotate on Monday
    interval=1,
    backupCount=8
)
file_handler.suffix = "_%Y-%m-%d.txt"
file_handler.namer = lambda name: name.replace(".txt_", "_")
file_handler.setLevel(logging.INFO)

# Console handler
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.WARNING)  # Only warnings and errors to console

# Formatter
formatter = logging.Formatter("%(asctime)s | %(levelname)-8s | %(message)s")
file_handler.setFormatter(formatter)
console_handler.setFormatter(formatter)

logger.addHandler(file_handler)
logger.addHandler(console_handler)

# ---------------------------------------------
# CONFIG
# ---------------------------------------------
# .env file should contain:
# SNOW_INSTANCE=your-instance.service-now.com
# SNOW_USER=your_username
# SNOW_PASS=your_password
# COCKPIT_PATH=C:\Users\cglynn\OneDrive - PHINIA\Data\EMEA_GOV\Cockpit\EMEA_Governance_Cockpit_2026.xlsx
# EUC_PATH=C:\Users\cglynn\OneDrive - PHINIA\My_Development_Projects\EMEA_GOV\Data\EUC_EOSL.xlsx
# SNOW_LOCATIONS_PATH=C:\Users\cglynn\OneDrive - PHINIA\Data\EMEA_GOV\SNOW_Exports\Current\PYTHON_EMEA_Locations.csv
#   (optional — only needed to override the default fallback path)

SNOW_INSTANCE = os.getenv("SNOW_INSTANCE")
SNOW_USER     = os.getenv("SNOW_USER")
SNOW_PASS     = os.getenv("SNOW_PASS")
SNOW_VERIFY_SSL = os.getenv("SNOW_VERIFY_SSL", "true").lower() == "true"
COCKPIT_PATH  = os.getenv("COCKPIT_PATH", "EMEA_Governance_Cockpit_2026.xlsx")
EUC_PATH      = os.getenv("EUC_PATH", "EUC_EOSL.xlsx")

# Fallback path for EMEA site list if cmn_location API call fails.
# Default: SNOW_Exports\Current\PYTHON_EMEA_Locations.csv relative to script directory.
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SNOW_LOCATIONS_PATH = os.getenv(
    "SNOW_LOCATIONS_PATH",
    os.path.join(_SCRIPT_DIR, "SNOW_Exports", "Current", "PYTHON_EMEA_Locations.csv")
)

BASE_URL = f"https://{SNOW_INSTANCE}/api/now/table"
AUTH     = (SNOW_USER, SNOW_PASS)
HEADERS  = {"Accept": "application/json", "Content-Type": "application/json"}

# Disable SSL warnings if verification is turned off
if not SNOW_VERIFY_SSL:
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

TODAY = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)

# SLA targets in hours — used if sla_target field absent from SNOW
SLA_DEFAULTS = {"1": 4, "2": 8, "3": 72, "4": 336}

# Non-physical / cloud / external entries to exclude from all site lists.
# These appear in the SNOW cmn_location EMEA region but are not physical governance sites.
# Blonie — manufacturing site closed (Mar 2026); may remain u_active in cmn_location.
LOCATION_EXCLUSIONS = [
    "Azure Europe West",
    "PHINIA Azure West Europe",
    "PHINIA Azure North Europe",
    "DXC Service Desk Budapest",
    "Abingdon - Cannon External Company",
    "Amsterdam Data Center NEW",
    "Amsterdam CoLocation",
    "Blonie - Poland",
]

# EMEA physical site lists — populated at runtime by fetch_emea_sites() in main().
# Do not edit these directly; update LOCATION_EXCLUSIONS above to control scope.
EMEA_SITES: list[str] = []            # u_site_name display names (Python-side filter)
EMEA_LOCATION_FILTER: str = ""        # location.nameIN<ids> (SNOW-side filter for sc_task/problem)

# LogicMonitor/Integration caller — exclude from incident metrics (monitoring/event tickets)
CALLER_EXCLUDE_SYS_ID = "24f555b8c387f6d41edf787dc00131a6"


# ---------------------------------------------
# SNOW API HELPERS
# ---------------------------------------------

def snow_query(table: str, query: str, fields: list[str],
               limit: int = 10000, retry: bool = True) -> pd.DataFrame:
    """
    Query a SNOW table via REST API. Returns a DataFrame.
    Handles pagination automatically if record count exceeds limit.
    Includes error handling and automatic retry on timeout.
    """
    params = {
        "sysparm_query":  query,
        "sysparm_fields": ",".join(fields),
        "sysparm_limit":  limit,
        "sysparm_offset": 0,
    }
    records = []

    # Debug: print query for first request
    if params["sysparm_offset"] == 0:
        logger.info(f"Query {table}: {query[:200]}")

    try:
        while True:
            try:
                # Add display_value=all to get both sys_id and display values
                params_with_display = params.copy()
                params_with_display["sysparm_display_value"] = "all"

                resp = requests.get(
                    f"{BASE_URL}/{table}",
                    auth=AUTH, headers=HEADERS, params=params_with_display,
                    timeout=30,
                    verify=SNOW_VERIFY_SSL
                )
                resp.raise_for_status()
                batch = resp.json().get("result", [])
                records.extend(batch)
                if len(batch) < limit:
                    break
                params["sysparm_offset"] += limit

                # Safety limit: don't fetch more than 50,000 records
                if len(records) >=50000:
                    logger.warning(f"Hit safety limit of 50,000 records for {table} - stopping fetch")
                    print(f"  WARNING: Fetched 50,000+ records from {table} - site filter may not be working")
                    break

            except requests.exceptions.Timeout:
                if retry:
                    logger.warning(f"Timeout querying {table} (offset {params['sysparm_offset']}), retrying once...")
                    print(f"  WARNING: Timeout - retrying...")
                    return snow_query(table, query, fields, limit, retry=False)
                else:
                    logger.error(f"Timeout querying {table} after retry - aborting")
                    print(f"\n  ERROR: SNOW API timeout - request took longer than 30 seconds")
                    print(f"     Check SNOW instance availability or network connection")
                    sys.exit(1)

    except requests.exceptions.ConnectionError as e:
        logger.error(f"Connection error querying {table}: {str(e)}")
        print(f"\n  ERROR: Cannot connect to SNOW instance: {SNOW_INSTANCE}")
        print(f"     Check network connection and instance URL in .env")
        sys.exit(1)

    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 401:
            logger.error(f"Authentication failed querying {table}: 401 Unauthorized")
            print(f"\n  ERROR: Authentication failed - check SNOW credentials in .env")
            print(f"     Verify SNOW_USER and SNOW_PASS are correct")
            sys.exit(1)
        elif e.response.status_code == 403:
            logger.error(f"Access forbidden querying {table}: 403 Forbidden")
            print(f"\n  ERROR: Access denied - check SNOW API access permissions")
            print(f"     User may not have read access to {table} table")
            sys.exit(1)
        else:
            logger.error(f"HTTP {e.response.status_code} querying {table}: {e.response.text}")
            print(f"\n  ERROR: SNOW API error: HTTP {e.response.status_code}")
            print(f"     {e.response.text[:200]}")
            sys.exit(1)

    except Exception as e:
        logger.error(f"Unexpected error querying {table}: {str(e)}")
        print(f"\n  ERROR: Unexpected error querying SNOW table {table}")
        print(f"     {str(e)}")
        sys.exit(1)

    df = pd.DataFrame(records)
    if df.empty:
        logger.warning(f"No records returned for {table} — check site filter")
        print(f"  WARNING: No records returned for {table} — check site filter")
        return df

    # Flatten dot-walked fields e.g. {"value": "...", "display_value": "..."}
    for col in df.columns:
        if df[col].dtype == object:
            sample = df[col].dropna().iloc[0] if not df[col].dropna().empty else None
            if isinstance(sample, dict) and "value" in sample:
                df[col] = df[col].apply(
                    lambda x: x.get("display_value") or x.get("value")
                    if isinstance(x, dict) else x
                )

    # Parse date fields
    for col in ["opened_at", "sys_updated_on", "resolved_at", "closed_at"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], utc=True, errors="coerce")

    logger.info(f"Successfully fetched {len(df)} records from {table}")

    # Debug: Print fields and sample data for first fetch
    if params["sysparm_offset"] == limit:  # After first batch
        logger.info(f"Fields in {table}: {list(df.columns)}")
        if "location" in df.columns:
            logger.info(f"Sample location values: {df['location'].dropna().unique()[:5].tolist()}")

    return df


# ---------------------------------------------
# EMEA SITE LIST — DYNAMIC FETCH
# ---------------------------------------------

def fetch_emea_sites() -> dict:
    """
    Fetch the authoritative EMEA physical site list from SNOW cmn_location.

    Primary source: REST API query against cmn_location filtered by u_region=EMEA and u_active=true.
    Fallback: CSV file at SNOW_LOCATIONS_PATH (PYTHON_EMEA_Locations.csv format).

    Excludes all entries in LOCATION_EXCLUSIONS (cloud, non-physical, external sites).

    Returns:
        dict with keys:
          - 'site_names'     : list[str]  — u_site_name display names (Python-side filter)
          - 'location_ids'   : str        — comma-separated numeric location names
                                            for SNOW-side location.nameIN filter
          - 'source'         : str        — 'api' or 'csv' (for logging)
          - 'count'          : int        — number of physical sites loaded
    """
    def _apply_exclusions(records: list[dict]) -> list[dict]:
        """Remove excluded entries from a list of {name, u_site_name} dicts."""
        return [
            r for r in records
            if r.get("u_site_name") not in LOCATION_EXCLUSIONS
        ]

    def _build_result(records: list[dict], source: str) -> dict:
        """Build the return dict from a filtered list of location records."""
        site_names   = [r["u_site_name"] for r in records if r.get("u_site_name")]
        location_ids = ",".join(r["name"] for r in records if r.get("name"))
        return {
            "site_names":   site_names,
            "location_ids": location_ids,
            "source":       source,
            "count":        len(site_names),
        }

    # --- Primary: SNOW REST API ---
    try:
        params = {
            "sysparm_query":  "u_region=EMEA^u_active=true",
            "sysparm_fields": "name,u_site_name,city,country",
            "sysparm_limit":  500,
        }
        resp = requests.get(
            f"{BASE_URL}/cmn_location",
            auth=AUTH, headers=HEADERS, params=params,
            verify=SNOW_VERIFY_SSL, timeout=30
        )
        resp.raise_for_status()
        raw = resp.json().get("result", [])

        if not raw:
            raise ValueError("cmn_location API returned 0 records — falling back to CSV")

        # Flatten dot-walked fields if present
        records = []
        for row in raw:
            name      = row.get("name", {})
            u_site    = row.get("u_site_name", {})
            records.append({
                "name":       name.get("value", name) if isinstance(name, dict) else str(name),
                "u_site_name": u_site.get("display_value") or u_site.get("value", u_site)
                               if isinstance(u_site, dict) else str(u_site),
            })

        records = _apply_exclusions(records)

        if not records:
            raise ValueError("No physical sites remain after exclusions — falling back to CSV")

        result = _build_result(records, "api")
        print(f"  [SITES] {result['count']} physical sites loaded from cmn_location (API)")
        logger.info(f"EMEA sites loaded from API — {result['count']} sites")
        return result

    except Exception as api_err:
        print(f"  [SITES] API fetch failed: {api_err}")
        print(f"  [SITES] Falling back to: {SNOW_LOCATIONS_PATH}")
        logger.warning(f"cmn_location API failed ({api_err}), falling back to CSV")

    # --- Fallback: CSV ---
    if not os.path.exists(SNOW_LOCATIONS_PATH):
        msg = (
            f"FATAL: cmn_location API failed and fallback CSV not found at:\n"
            f"  {SNOW_LOCATIONS_PATH}\n"
            f"Export PYTHON_EMEA_Locations.csv from SNOW or set SNOW_LOCATIONS_PATH in .env"
        )
        print(f"\n  {msg}")
        logger.error(msg)
        sys.exit(1)

    try:
        loc_df = pd.read_csv(SNOW_LOCATIONS_PATH, dtype=str)

        # Accept either 'u_site_name' or 'name' column as display name
        if "u_site_name" not in loc_df.columns and "name" in loc_df.columns:
            loc_df = loc_df.rename(columns={"name": "u_site_name"})

        # Filter to EMEA active rows if columns present
        if "u_region" in loc_df.columns:
            loc_df = loc_df[loc_df["u_region"].str.upper() == "EMEA"]
        if "u_active" in loc_df.columns:
            loc_df = loc_df[loc_df["u_active"].str.upper() == "TRUE"]

        # Build records list
        name_col = "name" if "name" in loc_df.columns else loc_df.columns[0]
        records = loc_df[[name_col, "u_site_name"]].rename(
            columns={name_col: "name"}
        ).dropna(subset=["u_site_name"]).to_dict("records")

        records = _apply_exclusions(records)

        if not records:
            print("  FATAL: No physical sites remain after exclusions in fallback CSV")
            logger.error("No physical sites in fallback CSV after exclusions")
            sys.exit(1)

        result = _build_result(records, "csv")
        print(f"  [SITES] {result['count']} physical sites loaded from fallback CSV")
        logger.info(f"EMEA sites loaded from CSV — {result['count']} sites | {SNOW_LOCATIONS_PATH}")
        return result

    except Exception as csv_err:
        msg = f"FATAL: Failed to read fallback CSV ({csv_err})"
        print(f"\n  {msg}")
        logger.error(msg)
        sys.exit(1)


# ---------------------------------------------
# DATA RETRIEVAL
# ---------------------------------------------

def fetch_incidents() -> pd.DataFrame:
    """Open incidents for all EMEA sites — batch-by-site.

    SNOW incident table ignores server-side location filters (nameIN, u_region).
    Querying one site at a time using location.u_site_name=<site> is the only
    reliable approach. Runs one query per EMEA site and concatenates results.
    Excludes LogicMonitor/Integration caller (monitoring/event tickets).
    Date filter: last 12 months (opened_at >= 365 days ago). Adjust timedelta if needed.
    """
    fields = [
        "number", "location.u_site_name", "priority",
        "opened_at", "sys_updated_on",
        "sla_target", "cmdb_ci", "problem_id",
        "state", "short_description"
    ]
    since_1yr = (TODAY - timedelta(days=365)).strftime("%Y-%m-%d %H:%M:%S")
    print(f"  Fetching open incidents (batch-by-site, {len(EMEA_SITES)} sites)...")
    dfs = []
    for i, site in enumerate(EMEA_SITES, 1):
        # 6=Resolved, 7=Closed, 8=Autoclosed
        query = (
            f"location.u_site_name={site}"
            f"^stateNOT IN6,7,8"
            f"^opened_at>={since_1yr}"
            f"^caller_id!={CALLER_EXCLUDE_SYS_ID}"
        )
        print(f"    [{i:>2}/{len(EMEA_SITES)}] {site[:45]}", end=" ", flush=True)
        df_site = snow_query("incident", query, fields)
        print(f"— {len(df_site)} records", flush=True)
        dfs.append(df_site)

    if not dfs or all(d.empty for d in dfs):
        logger.warning("fetch_incidents: no records returned for any EMEA site")
        return pd.DataFrame()

    df = pd.concat(dfs, ignore_index=True)

    if "location.u_site_name" in df.columns:
        df = df.rename(columns={"location.u_site_name": "location"})

    if "location" in df.columns:
        df = df[df["location"].isin(EMEA_SITES)]

    print(f"    Total open incidents: {len(df)}")
    logger.info(f"fetch_incidents complete — {len(df)} records across {len(EMEA_SITES)} sites")
    return df


def fetch_major_incident_history() -> pd.DataFrame:
    """P1/P2 incidents (open or closed) in the past 60 days — batch-by-site.

    Used for repeat MI check (calc_m5). Same batch-by-site approach as
    fetch_incidents() — SNOW location filters are ignored on incident table.
    Excludes LogicMonitor/Integration caller.
    """
    since = (TODAY - timedelta(days=60)).strftime("%Y-%m-%d %H:%M:%S")
    fields = [
        "number", "location.u_site_name", "priority",
        "opened_at", "resolved_at", "closed_at",
        "cmdb_ci", "problem_id", "state", "short_description"
    ]
    print(f"  Fetching MI history — 60 days (batch-by-site, {len(EMEA_SITES)} sites)...")
    dfs = []
    for i, site in enumerate(EMEA_SITES, 1):
        query = (
            f"location.u_site_name={site}"
            f"^priorityIN1,2"
            f"^opened_at>={since}"
            f"^caller_id!={CALLER_EXCLUDE_SYS_ID}"
        )
        print(f"    [{i:>2}/{len(EMEA_SITES)}] {site[:45]}", end=" ", flush=True)
        df_site = snow_query("incident", query, fields)
        print(f"— {len(df_site)} records", flush=True)
        dfs.append(df_site)

    if not dfs or all(d.empty for d in dfs):
        logger.warning("fetch_major_incident_history: no records returned for any EMEA site")
        return pd.DataFrame()

    df = pd.concat(dfs, ignore_index=True)

    if "location.u_site_name" in df.columns:
        df = df.rename(columns={"location.u_site_name": "location"})

    if "location" in df.columns:
        df = df[df["location"].isin(EMEA_SITES)]

    print(f"    Total MI history records: {len(df)}")
    logger.info(f"fetch_major_incident_history complete — {len(df)} records")
    return df


def fetch_catalogue_tasks() -> pd.DataFrame:
    """Open catalogue tasks for all 25 EMEA sites.

    sc_task does not populate location.u_site_name directly.
    Location is derived from the requesting user via:
      request_item.u_opened_on_behalf_of.location
    Values include a numeric site code prefix, e.g. "10610 - Warwick - United Kingdom".
    The prefix is stripped before matching against EMEA_SITES.
    """
    query  = EMEA_LOCATION_FILTER + "^stateNOT IN4,7"  # 4=Closed Complete, 7=Cancelled
    fields = [
        "number", "request_item.u_opened_on_behalf_of.location",
        "opened_at", "closed_at", "sys_updated_on",
        "state", "short_description",
    ]
    print("  Fetching open catalogue tasks...")
    df = snow_query("sc_task", query, fields)

    # Rename dot-walked field to simple column name
    loc_field = "request_item.u_opened_on_behalf_of.location"
    if loc_field in df.columns:
        df = df.rename(columns={loc_field: "location"})

    # Strip numeric site code prefix ("10610 - Warwick..." -> "Warwick...")
    if not df.empty and "location" in df.columns:
        df["location"] = df["location"].str.replace(r"^\d+ - ", "", regex=True)

    # Filter to EMEA sites only
    if not df.empty and "location" in df.columns:
        df = df[df["location"].isin(EMEA_SITES)]

    return df


def fetch_problems() -> pd.DataFrame:
    """Open EMEA problems where RCA is required.

    SNOW-side filters:
      - location IN EMEA site codes
      - u_rca_required = Yes
    Python-side:
      - problem_state != Closed  (state codes vary by instance)
    """
    query = EMEA_LOCATION_FILTER + "^u_rca_required=Yes"
    fields = [
        "number", "location.u_site_name", "location.u_region",
        "opened_at", "sys_updated_on",
        "u_root_cause", "u_root_cause_category",
        "problem_state", "short_description",
        "u_rca_required"
    ]
    print("  Fetching open problems...")
    df = snow_query("problem", query, fields)

    if df.empty:
        return df

    # Exclude closed problems (state codes vary by instance — filter in Python)
    if "problem_state" in df.columns:
        df = df[df["problem_state"].str.lower() != "closed"]

    # Rename dot-walked field to simple column name
    if "location.u_site_name" in df.columns:
        df = df.rename(columns={"location.u_site_name": "location"})

    return df


def _utc_snow_timestamp(dt: datetime) -> str:
    """Format a timezone-aware instant for SNOW encoded queries (UTC)."""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def fetch_historical_signal(as_of: datetime) -> dict:
    """
    Point-in-time SNOW datasets for EMEA scope (snapshot instant as_of, UTC).

    Returns dict with keys: incidents, mi_history, tasks, problems — same shapes
    as fetch_incidents / fetch_major_incident_history / fetch_catalogue_tasks /
    fetch_problems for downstream calc_m1–m6 (pass as_of into those calculators).
    """
    if as_of.tzinfo is None:
        as_of = as_of.replace(tzinfo=timezone.utc)
    as_of = as_of.astimezone(timezone.utc)
    as_of_str = _utc_snow_timestamp(as_of)
    since_1yr = _utc_snow_timestamp(as_of - timedelta(days=365))
    since_60 = _utc_snow_timestamp(as_of - timedelta(days=60))
    since_730 = _utc_snow_timestamp(as_of - timedelta(days=730))

    # --- Incidents (batch-by-site): open as of as_of ---
    inc_fields = [
        "number", "location.u_site_name", "priority",
        "opened_at", "sys_updated_on", "resolved_at",
        "sla_target", "cmdb_ci", "problem_id",
        "state", "short_description",
    ]
    print(f"  Historical incidents (as of {as_of_str} UTC, batch-by-site)...")
    dfs_inc = []
    for i, site in enumerate(EMEA_SITES, 1):
        query = (
            f"location.u_site_name={site}"
            f"^opened_at>={since_1yr}"
            f"^opened_at<{as_of_str}"
            f"^caller_id!={CALLER_EXCLUDE_SYS_ID}"
            f"^resolved_atISEMPTY^ORresolved_at>{as_of_str}"
        )
        print(f"    [{i:>2}/{len(EMEA_SITES)}] {site[:45]}", end=" ", flush=True)
        df_site = snow_query("incident", query, inc_fields)
        print(f"— {len(df_site)} records", flush=True)
        dfs_inc.append(df_site)

    if not dfs_inc or all(d.empty for d in dfs_inc):
        incidents = pd.DataFrame()
    else:
        incidents = pd.concat(dfs_inc, ignore_index=True)
        if "location.u_site_name" in incidents.columns:
            incidents = incidents.rename(columns={"location.u_site_name": "location"})
        if "location" in incidents.columns:
            incidents = incidents[incidents["location"].isin(EMEA_SITES)]

    # --- MI history (P1/P2, 60-day window ending at as_of) ---
    mi_fields = [
        "number", "location.u_site_name", "priority",
        "opened_at", "resolved_at", "closed_at",
        "cmdb_ci", "problem_id", "state", "short_description",
    ]
    print(f"  Historical MI history (60d window ending {as_of_str} UTC)...")
    dfs_mi = []
    for i, site in enumerate(EMEA_SITES, 1):
        query = (
            f"location.u_site_name={site}"
            f"^priorityIN1,2"
            f"^opened_at>={since_60}"
            f"^opened_at<{as_of_str}"
            f"^caller_id!={CALLER_EXCLUDE_SYS_ID}"
        )
        print(f"    [{i:>2}/{len(EMEA_SITES)}] {site[:45]}", end=" ", flush=True)
        df_site = snow_query("incident", query, mi_fields)
        print(f"— {len(df_site)} records", flush=True)
        dfs_mi.append(df_site)

    if not dfs_mi or all(d.empty for d in dfs_mi):
        mi_history = pd.DataFrame()
    else:
        mi_history = pd.concat(dfs_mi, ignore_index=True)
        if "location.u_site_name" in mi_history.columns:
            mi_history = mi_history.rename(columns={"location.u_site_name": "location"})
        if "location" in mi_history.columns:
            mi_history = mi_history[mi_history["location"].isin(EMEA_SITES)]

    # --- sc_task: EMEA location + point-in-time closure (closed_at applied in Python) ---
    task_fields = [
        "number", "request_item.u_opened_on_behalf_of.location",
        "opened_at", "closed_at", "sys_updated_on",
        "state", "short_description",
    ]
    print("  Historical catalogue tasks (sc_task)...")
    # Broad fetch: avoid encoded closed_at OR-clauses that can return 0 rows on some instances;
    # "open at snapshot" uses closed_at after fetch (see below).
    task_query = (
        EMEA_LOCATION_FILTER
        + f"^opened_at>={since_730}"
        + f"^opened_at<{as_of_str}"
    )
    tasks = snow_query("sc_task", task_query, task_fields)
    loc_field = "request_item.u_opened_on_behalf_of.location"
    if loc_field in tasks.columns:
        tasks = tasks.rename(columns={loc_field: "location"})
    if not tasks.empty and "location" in tasks.columns:
        tasks["location"] = tasks["location"].str.replace(r"^\d+ - ", "", regex=True)
        tasks = tasks[tasks["location"].isin(EMEA_SITES)]
    # Open-at-snapshot: opened_at < as_of AND (closed_at is NULL OR closed_at > as_of)
    if not tasks.empty:
        if "opened_at" in tasks.columns:
            tasks["opened_at"] = pd.to_datetime(tasks["opened_at"], utc=True, errors="coerce")
        if "closed_at" in tasks.columns:
            tasks["closed_at"] = pd.to_datetime(tasks["closed_at"], utc=True, errors="coerce")
            opened_before = tasks["opened_at"].notna() & (tasks["opened_at"] < as_of)
            still_open_at = tasks["closed_at"].isna() | (tasks["closed_at"] > as_of)
            tasks = tasks[opened_before & still_open_at]
        else:
            logger.warning(
                "Historical sc_task: closed_at not in API response — "
                "cannot apply point-in-time closure filter"
            )
            tasks = tasks[tasks["opened_at"].notna() & (tasks["opened_at"] < as_of)]

    # --- Problems: RCA required, open as of as_of (filter closed_at in Python) ---
    prob_fields = [
        "number", "location.u_site_name", "location.u_region",
        "opened_at", "sys_updated_on", "closed_at",
        "u_root_cause", "u_root_cause_category",
        "problem_state", "short_description",
        "u_rca_required",
    ]
    print("  Historical problems (RCA required)...")
    prob_query = EMEA_LOCATION_FILTER + "^u_rca_required=Yes" + f"^opened_at<{as_of_str}"
    problems = snow_query("problem", prob_query, prob_fields)
    if not problems.empty:
        if "closed_at" in problems.columns:
            problems = problems[
                problems["closed_at"].isna() | (problems["closed_at"] > as_of)
            ]
        else:
            logger.warning(
                "Historical problems: closed_at not in API response — "
                "filtering with problem_state!=closed (snapshot accuracy may be limited)"
            )
            if "problem_state" in problems.columns:
                problems = problems[problems["problem_state"].str.lower() != "closed"]
    if not problems.empty and "location.u_site_name" in problems.columns:
        problems = problems.rename(columns={"location.u_site_name": "location"})

    logger.info(
        f"fetch_historical_signal({as_of_str}): inc={len(incidents)} mi={len(mi_history)} "
        f"tasks={len(tasks)} prob={len(problems)}"
    )

    return {
        "incidents": incidents,
        "mi_history": mi_history,
        "tasks": tasks,
        "problems": problems,
    }


def load_from_csv(path: str) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Load SNOW data from a consolidated CSV file.
    Returns (incidents, mi_history, tasks, problems) matching the shape of fetch_*() outputs.
    """
    df_raw = pd.read_csv(path)
    total_count = len(df_raw)
    sys_classes = df_raw["sys_class_name"].dropna().unique().tolist() if "sys_class_name" in df_raw.columns else []
    print(f"  CSV: {total_count} total records, sys_class_name values: {sys_classes}")

    # Rename location columns to "location" (matches fetch_* column names for downstream metrics)
    if "location.u_site_name" in df_raw.columns:
        df_raw = df_raw.rename(columns={"location.u_site_name": "location"})
    elif "u_site_name" in df_raw.columns and "location" not in df_raw.columns:
        df_raw = df_raw.rename(columns={"u_site_name": "location"})
    if "request_item.u_opened_on_behalf_of.location" in df_raw.columns:
        ri_col = "request_item.u_opened_on_behalf_of.location"
        if "location" in df_raw.columns:
            df_raw["location"] = df_raw["location"].fillna(df_raw[ri_col])
        else:
            df_raw = df_raw.rename(columns={ri_col: "location"})

    # Site name normalisation: strip numeric prefix (e.g. "10610 - Warwick - United Kingdom")
    if "location" in df_raw.columns:
        df_raw["location"] = df_raw["location"].apply(
            lambda x: re.sub(r"^\d+\s*-\s*", "", str(x)).strip() if pd.notna(x) else x
        )

    # Parse date columns
    for col in ["opened_at", "sys_updated_on", "resolved_at", "closed_at"]:
        if col in df_raw.columns:
            df_raw[col] = pd.to_datetime(df_raw[col], utc=True, errors="coerce")

    # Split by sys_class_name and apply filters
    cutoff_60d = TODAY - timedelta(days=60)

    incidents = pd.DataFrame()
    if "incident" in df_raw.get("sys_class_name", pd.Series()).values:
        inc_raw = df_raw[df_raw["sys_class_name"] == "incident"].copy()
        if "state" in inc_raw.columns:
            inc_raw["state"] = pd.to_numeric(inc_raw["state"], errors="coerce")
            incidents = inc_raw[~inc_raw["state"].isin([6, 7])]
        else:
            incidents = inc_raw

    mi_history = pd.DataFrame()
    if "incident" in df_raw.get("sys_class_name", pd.Series()).values:
        mi_raw = df_raw[df_raw["sys_class_name"] == "incident"].copy()
        if "priority" in mi_raw.columns and "opened_at" in mi_raw.columns:
            mi_raw = mi_raw[mi_raw["priority"].astype(str).isin(["1", "2"])]
            mi_raw = mi_raw[mi_raw["opened_at"] >= cutoff_60d]
        mi_history = mi_raw

    tasks = pd.DataFrame()
    if "sc_task" in df_raw.get("sys_class_name", pd.Series()).values:
        task_raw = df_raw[df_raw["sys_class_name"] == "sc_task"].copy()
        if "state" in task_raw.columns:
            task_raw["state"] = pd.to_numeric(task_raw["state"], errors="coerce")
            tasks = task_raw[~task_raw["state"].isin([4, 7])]
        else:
            tasks = task_raw

    problems = pd.DataFrame()
    if "problem" in df_raw.get("sys_class_name", pd.Series()).values:
        prob_raw = df_raw[df_raw["sys_class_name"] == "problem"].copy()
        if "state" in prob_raw.columns:
            prob_raw["state"] = pd.to_numeric(prob_raw["state"], errors="coerce")
            problems = prob_raw[~prob_raw["state"].isin([4, 7])]
        else:
            problems = prob_raw

        # Root cause normalisation (match fetch_problems: ensure u_root_cause for calc_m6)
        if "u_root_cause" in problems.columns and "root_cause" not in problems.columns:
            problems["root_cause"] = problems["u_root_cause"]
        elif "u_root_cause" in problems.columns and "root_cause" in problems.columns:
            problems["root_cause"] = problems["root_cause"].fillna(problems["u_root_cause"])
        if "root_cause" in problems.columns and "u_root_cause" not in problems.columns:
            problems["u_root_cause"] = problems["root_cause"]
        if "u_root_cause" not in problems.columns:
            problems["u_root_cause"] = pd.NA

    # Filter to EMEA sites for all DataFrames
    for dfs in [incidents, mi_history, tasks, problems]:
        if not dfs.empty and "location" in dfs.columns:
            dfs.drop(dfs[~dfs["location"].isin(EMEA_SITES)].index, inplace=True)

    print(f"  CSV mode — loaded from: {path}")
    print(f"  Incidents (open):    {len(incidents):>5} records")
    print(f"  MI history (60d):    {len(mi_history):>5} records")
    print(f"  Catalogue tasks:     {len(tasks):>5} records")
    print(f"  Problems:            {len(problems):>5} records")

    return (incidents, mi_history, tasks, problems)


def test_incident_fetch(limit: int = 1000, batch_sites: int = 3) -> None:
    """
    Test different incident fetch strategies to validate filters and data volume.
    Run with: python emea_gov_refresh.py --test-fetch

    limit: Max records per query for quicker tests (default 1000). Use 50000 for full run.
    batch_sites: Number of sites to test for batch-by-site (default 3).
    """
    global EMEA_SITES, EMEA_LOCATION_FILTER
    if not EMEA_SITES:
        print("  Initialising EMEA site list for test...")
        sites = fetch_emea_sites()
        EMEA_SITES = sites["site_names"]
        EMEA_LOCATION_FILTER = "location.nameIN" + sites["location_ids"]
        print(f"  {sites['count']} EMEA sites loaded (source: {sites['source']})\n")

    incident_fields = [
        "number", "location.u_site_name", "priority",
        "opened_at", "sys_updated_on", "caller_id", "contact_type", "incident_state",
        "sla_target", "cmdb_ci", "problem_id", "state", "short_description"
    ]
    loc_col = "location.u_site_name"

    def _postprocess(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        if loc_col in df.columns:
            df = df.rename(columns={loc_col: "location"})
        if "location" in df.columns:
            df = df[df["location"].isin(EMEA_SITES)]
        return df

    strategies = [
        ("Current (state only)", "stateNOT IN6,7,8"),
        ("+ caller exclude", f"stateNOT IN6,7,8^caller_id!={CALLER_EXCLUDE_SYS_ID}"),
        ("+ location.nameIN", f"{EMEA_LOCATION_FILTER}^stateNOT IN6,7,8^caller_id!={CALLER_EXCLUDE_SYS_ID}"),
        ("+ location.u_region=EMEA", f"location.u_region=EMEA^stateNOT IN6,7,8^caller_id!={CALLER_EXCLUDE_SYS_ID}"),
    ]

    print(f"\n{'='*60}")
    print("INCIDENT FETCH TEST")
    print(f"limit={limit} per query | batch_sites={batch_sites}")
    print(f"{'='*60}\n")

    for i, (name, query) in enumerate(strategies, 1):
        try:
            print(f"  [{i}/{len(strategies)}] {name}...", end=" ", flush=True)
            t0 = time.time()
            df = snow_query("incident", query, incident_fields, limit=limit)
            raw = len(df)
            df = _postprocess(df)
            emea = len(df)
            elapsed = time.time() - t0
            hit_cap = " (HIT LIMIT)" if raw >= limit else ""
            print(f"Raw: {raw:,}{hit_cap}  EMEA: {emea:,}  {elapsed:.1f}s", flush=True)
        except Exception as e:
            print(f"  {name}: ERROR — {e}")

    # Batch-by-site (sample)
    print(f"\n  Batch-by-site (first {batch_sites} sites)", flush=True)
    try:
        t0 = time.time()
        dfs = []
        for j, site in enumerate(EMEA_SITES[:batch_sites], 1):
            print(f"    Site {j}/{batch_sites}: {site[:40]}...", end=" ", flush=True)
            q = f"location.u_site_name={site}^stateNOT IN6,7,8^caller_id!={CALLER_EXCLUDE_SYS_ID}"
            d = snow_query("incident", q, incident_fields, limit=limit)
            print(f"{len(d):,} records", flush=True)
            if not d.empty and loc_col in d.columns:
                d = d.rename(columns={loc_col: "location"})
            dfs.append(d)
        df_batch = pd.concat(dfs, ignore_index=True)
        df_batch = df_batch[df_batch["location"].isin(EMEA_SITES)] if "location" in df_batch.columns else df_batch
        elapsed = time.time() - t0
        print(f"    Total: {len(df_batch):,}  {elapsed:.1f}s")
    except Exception as e:
        print(f"    ERROR — {e}")

    print(f"\n{'='*60}")
    print("If '+ location.nameIN' or '+ location.u_region' show low Raw count,")
    print("  SNOW-side filter works — prefer that in production.")
    print("If all show Raw=limit, implement batch-by-site.")
    print(f"{'='*60}\n")


def fetch_euc_assets(euc_path: str) -> pd.DataFrame:
    """
    Read EUC/EOSL device replacement data from Power BI export Excel file.
    Returns DataFrame with columns: site, target_date_label, device_count
    Source: Export sheet in EUC_EOSL.xlsx (device-level records)
    """
    try:
        print("  Reading EUC/EOSL data from Power BI export...")

        if not os.path.exists(euc_path):
            logger.warning(f"EUC file not found: {euc_path}")
            print(f"  WARNING: EUC/EOSL file not found - skipping Physics Block 1")
            return pd.DataFrame()

        # Read the Export sheet (Power BI export format)
        wb = openpyxl.load_workbook(euc_path, data_only=True)

        if "Export" not in wb.sheetnames:
            logger.warning(f"'Export' sheet not found in {euc_path}")
            print(f"  WARNING: Expected sheet 'Export' not found - skipping Physics Block 1")
            print(f"     Available sheets: {', '.join(wb.sheetnames)}")
            return pd.DataFrame()

        ws = wb["Export"]

        # Read device-level data into pandas DataFrame
        data_rows = []
        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):  # Skip completely empty rows
                data_rows.append(row)

        df_devices = pd.DataFrame(data_rows, columns=headers)

        # Filter to devices that need replacement
        if 'Action to take' not in df_devices.columns or 'Site Location' not in df_devices.columns:
            logger.warning(f"Required columns missing in EUC file: {euc_path}")
            print(f"  WARNING: Expected columns 'Action to take' and 'Site Location' not found")
            return pd.DataFrame()

        # Filter to replacement devices only
        replacement_devices = df_devices[
            df_devices['Action to take'].astype(str).str.contains('Replace', case=False, na=False) |
            df_devices['Action to take'].astype(str).str.contains('Urgent', case=False, na=False)
        ].copy()

        if replacement_devices.empty:
            logger.warning(f"No replacement devices found in {euc_path}")
            print(f"  WARNING: No device replacement data found in EUC file")
            return pd.DataFrame()

        # Extract target date from "Action to take" field
        def extract_target_date(action):
            import re
            action_str = str(action)

            if "Urgent" in action_str:
                return "Urgent"

            # Match date patterns like "14/10/2025" or "11/11/2026"
            match = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', action_str)
            if match:
                day, month, year = match.groups()
                month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
                month_name = month_names[int(month) - 1]
                return f"{month_name} {year}"

            return "Unknown"

        replacement_devices['target_date_label'] = replacement_devices['Action to take'].apply(extract_target_date)

        # Group by site and target date
        summary = replacement_devices.groupby(['Site Location', 'target_date_label']).size().reset_index(name='device_count')
        summary = summary.rename(columns={'Site Location': 'site'})

        # Filter out "Unknown" dates
        summary = summary[summary['target_date_label'] != 'Unknown']

        if summary.empty:
            logger.warning(f"No EUC/EOSL data found in {euc_path}")
            print(f"  WARNING: No device replacement data found in EUC file")
            return summary

        logger.info(f"Successfully loaded {summary['device_count'].sum()} EUC/EOSL device records from {len(summary['site'].unique())} sites")
        print(f"  EUC/EOSL data:       {summary['device_count'].sum():>5} devices across {len(summary['site'].unique())} sites")

        return summary

    except PermissionError:
        logger.warning(f"Permission denied - EUC file is open: {euc_path}")
        print(f"  WARNING: Cannot read EUC file - it is currently open")
        print(f"     Close {euc_path} and re-run")
        return pd.DataFrame()

    except Exception as e:
        logger.error(f"Error reading EUC file: {str(e)}")
        print(f"  WARNING: Error reading EUC/EOSL file: {str(e)}")
        return pd.DataFrame()


# ---------------------------------------------
# METRIC CALCULATIONS
# ---------------------------------------------

def age_days(
    df: pd.DataFrame,
    date_col: str = "opened_at",
    as_of: datetime | None = None,
) -> pd.Series:
    """Return age in days from date_col to as_of (or TODAY if as_of is None)."""
    end = as_of if as_of is not None else TODAY
    if getattr(end, "tzinfo", None) is None:
        end = end.replace(tzinfo=timezone.utc)
    return (end - df[date_col]).dt.total_seconds() / 86400


def calc_m1_incident_aging(
    incidents: pd.DataFrame, as_of: datetime | None = None
) -> dict:
    """Metric 1: % open incidents aged <= 10 days."""
    if incidents.empty:
        return {"value": None, "note": "No data"}
    incidents = incidents.copy()
    incidents["age_days"] = age_days(incidents, "opened_at", as_of)
    compliant = (incidents["age_days"] <= 10).sum()
    total     = len(incidents)
    pct       = round(compliant / total * 100, 1) if total else 0
    return {
        "value": pct,
        "note":  f"{compliant}/{total} incidents aged <=10 days"
    }


def calc_m2_catalogue_aging(
    tasks: pd.DataFrame, as_of: datetime | None = None
) -> dict:
    """Metric 2: % open catalogue tasks aged <= 30 days."""
    if tasks.empty:
        return {"value": None, "note": "No data"}
    tasks = tasks.copy()
    tasks["opened_at"] = pd.to_datetime(tasks["opened_at"], utc=True, errors="coerce")
    if as_of is not None and "closed_at" in tasks.columns:
        tasks["closed_at"] = pd.to_datetime(tasks["closed_at"], utc=True, errors="coerce")
        end = as_of
        if getattr(end, "tzinfo", None) is None:
            end = end.replace(tzinfo=timezone.utc)
        else:
            end = end.astimezone(timezone.utc)
        tasks = tasks[tasks["closed_at"].isna() | (tasks["closed_at"] > end)]
    if tasks.empty:
        return {"value": None, "note": "No data"}
    tasks["age_days"] = age_days(tasks, "opened_at", as_of)
    compliant = (tasks["age_days"] <= 30).sum()
    total     = len(tasks)
    pct       = round(compliant / total * 100, 1) if total else 0
    return {
        "value": pct,
        "note":  f"{compliant}/{total} tasks aged <=30 days"
    }


def calc_m3_sla_x2(
    incidents: pd.DataFrame, as_of: datetime | None = None
) -> dict:
    """Metric 3: Count of open incidents exceeding SLA x2, split by priority."""
    if incidents.empty:
        return {"value": 0, "note": "No data", "p12_count": 0, "p34_count": 0}
    inc = incidents.copy()
    inc["age_days"] = age_days(inc, "opened_at", as_of)

    # SLA target in days per record
    if "sla_target" in inc.columns and inc["sla_target"].notna().any():
        inc["sla_days"] = pd.to_numeric(inc["sla_target"], errors="coerce") / 24
    else:
        inc["sla_days"] = inc["priority"].map(
            {k: v / 24 for k, v in SLA_DEFAULTS.items()}
        )

    inc["sla_x2"] = inc["age_days"] > (inc["sla_days"] * 2)
    breached = inc[inc["sla_x2"]]
    p12 = breached[breached["priority"].isin(["1", "2"])].shape[0]
    p34 = breached[breached["priority"].isin(["3", "4"])].shape[0]
    flag = " WARNING: P1/P2 PRESENT" if p12 > 0 else ""
    return {
        "value":    p12 + p34,
        "note":     f"P1/P2: {p12} | P3/P4: {p34}{flag}",
        "p12_count": p12,
        "p34_count": p34,
    }


def calc_m4_no_movement(
    incidents: pd.DataFrame,
    tasks: pd.DataFrame,
    as_of: datetime | None = None,
) -> dict:
    """Metric 4: Count of incidents with no update for >= 14 days (incidents only).

    Catalogue tasks are excluded — their health is captured by catalogue_aging (M2).
    no_movement is a lead indicator for incident SLA breaches; catalogue_aging is
    the lag indicator for catalogue backlog age.

    Historical snapshots: only rows with sys_updated_on <= as_of are considered;
    staleness is as_of - sys_updated_on. True point-in-time “no updates in the 14
    days before as_of” is not fully reconstructable from sys_updated_on alone
    without audit/history (sys_updated_on reflects last update as of today’s record).
    """
    if incidents.empty:
        return {"value": 0, "note": "No data"}

    end = as_of if as_of is not None else TODAY
    if getattr(end, "tzinfo", None) is None:
        end = end.replace(tzinfo=timezone.utc)

    inc = incidents.copy()
    inc = inc[inc["sys_updated_on"].notna()]
    inc = inc[inc["sys_updated_on"] <= end]
    inc["stale_days"] = (end - inc["sys_updated_on"]).dt.total_seconds() / 86400
    stale = inc[inc["stale_days"] >= 14]

    site_counts = {}
    if "location" in stale.columns:
        for site, cnt in stale["location"].value_counts().items():
            site_counts[site] = site_counts.get(site, 0) + cnt

    top5 = sorted(site_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    top5_str = ", ".join(f"{s}:{c}" for s, c in top5) if top5 else "none"
    return {
        "value": len(stale),
        "note":  f"Top sites: {top5_str}"
    }


def calc_m5_repeat_mi(mi_history: pd.DataFrame) -> dict:
    """Metric 5: Repeat MIs for same CI or site within any 30-day window."""
    if mi_history.empty:
        return {"value": 0, "note": "No data", "watch_count": 0, "breach_count": 0}

    mi = mi_history.copy().sort_values("opened_at")
    watch_items   = []
    breach_items  = []

    for (site, ci), grp in mi.groupby(["location", "cmdb_ci"]):
        dates = grp["opened_at"].tolist()
        for i in range(len(dates)):
            for j in range(i + 1, len(dates)):
                delta = abs((dates[j] - dates[i]).days)
                if delta <= 30:
                    no_problem = grp["problem_id"].isna().any() or \
                                 (grp["problem_id"] == "").any()
                    entry = f"{site} / CI: {ci or 'unknown'}"
                    if no_problem:
                        watch_items.append(entry)
                    else:
                        breach_items.append(entry)

    watch_items  = list(set(watch_items))
    breach_items = list(set(breach_items))

    note_parts = []
    if breach_items:
        note_parts.append("BREACH: " + "; ".join(breach_items[:3]))
    if watch_items:
        note_parts.append("WATCH: " + "; ".join(watch_items[:3]))

    return {
        "value":        len(watch_items) + len(breach_items),
        "note":         " | ".join(note_parts) if note_parts else "No repeats",
        "watch_count":  len(watch_items),
        "breach_count": len(breach_items),
    }


def calc_m6_problems_no_rca(
    problems: pd.DataFrame, as_of: datetime | None = None
) -> dict:
    """Metric 6: Open EMEA problems (RCA required) with no root cause and age > 30 days.

    Filters applied upstream in fetch_problems():
      - problem_state != Closed
      - u_rca_required = Yes
      - location is EMEA site or EMEA region
    Here we further filter to age > 30 days and u_root_cause blank.
    """
    if problems.empty:
        return {"value": 0, "note": "No data", "watch_count": 0, "breach_count": 0}

    p = problems.copy()
    p["age_days"] = age_days(p, "opened_at", as_of)

    # Age > 30 days
    aged = p[p["age_days"] > 30].copy()

    # No root cause populated
    no_rca = aged[
        aged["u_root_cause"].isna() | (aged["u_root_cause"] == "")
    ]

    watch  = no_rca[(no_rca["age_days"] > 30) & (no_rca["age_days"] <= 60)]
    breach = no_rca[no_rca["age_days"] > 60]

    return {
        "value":        len(no_rca),
        "note":         f"Watch (30-60d): {len(watch)} | Breach (60d+): {len(breach)}",
        "watch_count":  len(watch),
        "breach_count": len(breach),
    }


def calc_physics_block1_euc(euc_df: pd.DataFrame) -> dict:
    """
    Calculate Physics Block 1 metrics from EUC/EOSL device replacement data.
    Returns dict with: total_units, overdue, remaining, weeks_remaining,
    required_weekly_burn, q4_adjusted_burn.
    """
    if euc_df.empty:
        return {
            "total_units": 0,
            "overdue": 0,
            "remaining": 0,
            "weeks_remaining": 0,
            "required_weekly_burn": 0,
            "q4_adjusted_burn": 0,
            "note": "No EUC data"
        }

    # Total devices to replace
    total_units = euc_df["device_count"].sum()

    # Overdue: devices with target dates in the past (Oct 2025 or earlier)
    overdue_labels = ["Oct 2025", "Sep 2025", "Aug 2025", "Jul 2025", "Jun 2025",
                      "May 2025", "Apr 2025", "Mar 2025", "Feb 2025", "Jan 2025",
                      "Dec 2024", "Nov 2024", "Oct 2024"]  # Add more historical dates if needed
    overdue = euc_df[euc_df["target_date_label"].isin(overdue_labels)]["device_count"].sum()

    # Remaining by Nov 2026
    remaining_labels = ["Nov 2026", "Dec 2026"]  # Include Dec if exists
    remaining = euc_df[euc_df["target_date_label"].isin(remaining_labels)]["device_count"].sum()

    # Calculate weeks remaining to 31 Dec 2026 (to match cockpit)
    target_date = datetime(2026, 12, 31, tzinfo=timezone.utc)
    weeks_remaining = max(1, (target_date - TODAY).days / 7)  # At least 1 week to avoid division by zero

    # Required weekly burn
    required_weekly_burn = round(remaining / weeks_remaining, 1) if weeks_remaining > 0 else 0

    # Q4 adjusted burn (70% capacity from 1 Oct due to seasonal factors)
    # Check if we're currently in Q4 (Oct-Dec)
    if TODAY.month >= 10:
        q4_adjusted_burn = round(required_weekly_burn * 0.7, 1)
    else:
        q4_adjusted_burn = required_weekly_burn

    return {
        "total_units": int(total_units),
        "overdue": int(overdue),
        "remaining": int(remaining),
        "weeks_remaining": round(weeks_remaining, 1),
        "required_weekly_burn": required_weekly_burn,
        "q4_adjusted_burn": q4_adjusted_burn,
        "note": f"Total: {int(total_units)} | Overdue: {int(overdue)} | Remaining by Nov 2026: {int(remaining)}"
    }


# ---------------------------------------------
# TRIGGER STATUS EVALUATION
# ---------------------------------------------

def evaluate_trigger(metric_name: str, current: float,
                     prev_4w_avg: float, prev_week: float,
                     extra: dict = None) -> str:
    """
    Returns GREEN / WATCH / BREACHED based on trigger rules.
    extra: dict of additional context (p12_count, watch_count, breach_count etc.)
    """
    extra = extra or {}

    if metric_name == "incident_aging":
        if current is not None and current < 90 and prev_week is not None and prev_week < 90:
            return "BREACHED"
        if current is not None and current < 90:
            return "WATCH"
        return "GREEN"

    elif metric_name == "catalogue_aging":
        if current is not None and current < 90 and prev_week is not None and prev_week < 90:
            return "BREACHED"
        if current is not None and current < 90:
            return "WATCH"
        return "GREEN"

    elif metric_name == "no_movement":
        if current and current > 0 and prev_week and current > prev_week:
            return "BREACHED"  # increasing for >=2 weeks — simplification: upward trend
        if current and current > 0:
            return "WATCH"
        return "GREEN"

    elif metric_name == "repeat_mi":
        if extra.get("breach_count", 0) > 0:
            return "BREACHED"
        if extra.get("watch_count", 0) > 0:
            return "WATCH"
        return "GREEN"

    elif metric_name == "problems_no_rca":
        if extra.get("breach_count", 0) > 0:
            return "BREACHED"
        if extra.get("watch_count", 0) > 0:
            return "WATCH"
        return "GREEN"

    elif metric_name == "sla_x2":
        if extra.get("p12_count", 0) > 0:
            return "BREACHED"
        if current and current > 0 and prev_week and current > prev_week:
            return "BREACHED"
        if current and current > 0:
            return "WATCH"
        return "GREEN"

    return "PENDING"


# ---------------------------------------------
# EXCEL WRITE
# ---------------------------------------------

# Trigger status colour fills
FILLS = {
    "GREEN":   PatternFill("solid", fgColor="00B050"),
    "WATCH":   PatternFill("solid", fgColor="FFC000"),
    "BREACHED":PatternFill("solid", fgColor="FF0000"),
    "PENDING": PatternFill("solid", fgColor="D9D9D9"),
}
FONTS = {
    "GREEN":    Font(bold=True, color="FFFFFF"),
    "WATCH":    Font(bold=True, color="000000"),
    "BREACHED": Font(bold=True, color="FFFFFF"),
    "PENDING":  Font(bold=False, color="000000"),
}


def set_trigger_cell(ws, cell_ref: str, status: str):
    cell = ws[cell_ref]
    cell.value = status
    cell.fill  = FILLS.get(status, FILLS["PENDING"])
    cell.font  = FONTS.get(status, FONTS["PENDING"])


def shift_physics_trends(ws_physics, row: int, new_value: float):
    """
    Shift Wk1->Wk2, Wk2->Wk3, Wk3->Wk4 in Physics_Engine Block 4.
    Then write new_value into Wk1.
    Column layout assumed: A=Metric, B=Wk1, C=Wk2, D=Wk3, E=Wk4
    (adjust col letters to match your actual workbook layout)
    """
    cols = ["B", "C", "D", "E"]  # Wk1, Wk2, Wk3, Wk4
    # Shift right to left: E=D, D=C, C=B
    for i in range(len(cols) - 1, 0, -1):
        src = ws_physics[f"{cols[i-1]}{row}"].value
        ws_physics[f"{cols[i]}{row}"] = src
    ws_physics[f"{cols[0]}{row}"] = new_value  # new value into Wk1


def update_enterprise_panel_euc(ws_enterprise, euc_metrics: dict):
    """
    Write EUC stream 1 (Tech Refresh) to Enterprise_Panel row 4 — fixed layout (25 Mar 2026).
    C4 remaining units (int), E4 weeks to 31 Dec 2026, F4 required weekly burn, H4 physics status.
    G4 is manual (Actual 4W Rolling Burn) — read only, never overwritten.
    """
    # Stream 1 fixed refs — no label row search
    CELL_UNITS_REMAINING = "C4"
    CELL_WEEKS_REMAINING = "E4"
    CELL_REQUIRED_BURN = "F4"
    CELL_ACTUAL_BURN = "G4"  # read-only
    CELL_PHYSICS_STATUS = "H4"

    units_remaining = int(euc_metrics.get("remaining", 0) or 0)

    end_date = datetime(2026, 12, 31, tzinfo=timezone.utc).date()
    today_date = TODAY.date()
    days_to_end = (end_date - today_date).days
    weeks_remaining = round(days_to_end / 7, 1)

    if weeks_remaining and weeks_remaining > 0:
        required_burn = round(units_remaining / weeks_remaining, 1)
    else:
        required_burn = 0.0

    raw_actual = ws_enterprise[CELL_ACTUAL_BURN].value
    try:
        actual_burn = float(raw_actual) if raw_actual is not None and str(raw_actual).strip() != "" else 0.0
    except (TypeError, ValueError):
        actual_burn = 0.0

    if actual_burn >= required_burn:
        physics_status = "ON TRAJECTORY"
    elif actual_burn >= required_burn * 0.8:
        physics_status = "WATCH"
    else:
        physics_status = "BREACHED"

    ws_enterprise[CELL_UNITS_REMAINING].value = units_remaining
    ws_enterprise[CELL_WEEKS_REMAINING].value = weeks_remaining
    ws_enterprise[CELL_REQUIRED_BURN].value = required_burn
    ws_enterprise[CELL_PHYSICS_STATUS].value = physics_status

    logger.info(
        f"Enterprise_Panel EUC (row 4): remaining={units_remaining}, weeks={weeks_remaining}, "
        f"required_burn={required_burn}, actual_burn={actual_burn}, status={physics_status}"
    )


def save_enterprise_panel_euc_after_cockpit(cockpit_path: str, euc_metrics: dict | None) -> None:
    """
    Second pass on the cockpit workbook: Enterprise_Panel EUC row 4 only.
    If EUC_EOSL source file is missing, print a warning and return without raising.
    """
    if not os.path.exists(EUC_PATH):
        print(f"  WARNING: EUC_EOSL.xlsx not found ({EUC_PATH}) — skipping Enterprise_Panel EUC update")
        logger.warning(f"Enterprise_Panel EUC skipped — EUC file not found: {EUC_PATH}")
        return
    if euc_metrics is None:
        return
    try:
        wb = openpyxl.load_workbook(cockpit_path)
    except FileNotFoundError:
        print(f"  WARNING: Cockpit not found for Enterprise_Panel pass — {cockpit_path}")
        logger.warning(f"Enterprise_Panel EUC skipped — cockpit missing: {cockpit_path}")
        return
    except PermissionError:
        print(f"  WARNING: Cockpit open — could not update Enterprise_Panel EUC ({cockpit_path})")
        logger.warning(f"Enterprise_Panel EUC skipped — permission denied: {cockpit_path}")
        return
    except Exception as e:
        print(f"  WARNING: Could not load cockpit for Enterprise_Panel EUC: {e}")
        logger.warning(f"Enterprise_Panel EUC skipped — load error: {e}")
        return

    if "Enterprise_Panel" not in wb.sheetnames:
        print("  WARNING: Enterprise_Panel sheet missing — skipping EUC row update")
        logger.warning("Enterprise_Panel EUC skipped — sheet not in workbook")
        return

    try:
        update_enterprise_panel_euc(wb["Enterprise_Panel"], euc_metrics)
        wb.save(cockpit_path)
        logger.info(f"Enterprise_Panel EUC saved to {cockpit_path}")
    except PermissionError:
        print(f"  WARNING: Cannot save cockpit after Enterprise_Panel EUC update — file may be open")
        logger.warning(f"Enterprise_Panel EUC save failed — permission denied: {cockpit_path}")
    except Exception as e:
        print(f"  WARNING: Enterprise_Panel EUC update/save failed: {e}")
        logger.warning(f"Enterprise_Panel EUC save error: {e}")


def update_physics_block1(ws_physics, euc_metrics: dict):
    """
    Write EUC/EOSL device replacement metrics to Physics_Engine Block 1.
    Calibrated from EMEA_Governance_Cockpit_2026.xlsx on 2026-03-07
    """
    # Physics Block 1 cell references (calibrated to actual cockpit layout)
    BLOCK1_CELLS = {
        "total_units":           "B2",   # Total EOSL Units in Scope
        "units_replaced":        "B3",   # Units Remediated to Date (not auto-updated - manual input)
        "remaining":             "B5",   # Units Remaining
        "weeks_remaining":       "B6",   # Weeks to 31 Dec 2026
        "required_weekly_burn":  "B7",   # Required Weekly Burn
    }

    try:
        # Write calculated values (skip units_replaced - manually maintained)
        ws_physics[BLOCK1_CELLS["total_units"]].value = euc_metrics.get("total_units")
        # ws_physics[BLOCK1_CELLS["units_replaced"]] - DO NOT OVERWRITE (manual)
        ws_physics[BLOCK1_CELLS["remaining"]].value = euc_metrics.get("remaining")
        ws_physics[BLOCK1_CELLS["weeks_remaining"]].value = euc_metrics.get("weeks_remaining")
        ws_physics[BLOCK1_CELLS["required_weekly_burn"]].value = euc_metrics.get("required_weekly_burn")

        logger.info(f"Physics Block 1 updated - Total: {euc_metrics.get('total_units')}, "
                   f"Remaining: {euc_metrics.get('remaining')}, "
                   f"Weekly burn: {euc_metrics.get('required_weekly_burn')}")

    except Exception as e:
        logger.warning(f"Error updating Physics Block 1: {str(e)}")
        print(f"  WARNING: Could not update Physics Block 1 - check cell references")


METRIC_DISPLAY_NAMES = {
    "incident_aging":   "Incident Aging",
    "catalogue_aging":  "Catalogue Aging",
    "sla_x2":           "SLA x2",
    "no_movement":      "No Movement",
    "repeat_mi":        "Repeat MI",
    "problems_no_rca":  "Problems Without RCA",
}

TRIGGER_TYPE_MAP = {
    "incident_aging":   "Threshold — Consecutive Week",
    "catalogue_aging":  "Threshold — Consecutive Week",
    "sla_x2":           "Stability — P1/P2 in SLA x2 band",
    "no_movement":      "Stagnation — Count Increasing",
    "repeat_mi":        "Stability — No Problem Record",
    "problems_no_rca":  "Threshold — Age Band (60+ days)",
}

RULE_ID_MAP = {
    "incident_aging":   "OPS-T2",
    "catalogue_aging":  "OPS-T3",
    "sla_x2":           "OPS-T4",
    "no_movement":      "OPS-T7",
    "repeat_mi":        "OPS-T9",
    "problems_no_rca":  "OPS-T11",
}

ESCALATION_OWNER_MAP = {
    "incident_aging":   "IT Ops Managers — Proyer / Cazan / Gauthier / S K / Glynn",
    "catalogue_aging":  "IT Ops Managers — Proyer / Cazan / Gauthier / S K / Glynn",
    "sla_x2":           "IT Ops Managers — P1/P2: same day. P3/P4: per site grouping",
    "no_movement":      "IT Ops Managers — per site grouping",
    "repeat_mi":        "IT Ops Manager for breaching site — same day",
    "problems_no_rca":  "IT Ops Managers — per site grouping",
}

ESCALATION_RECIPIENTS = {
    "incident_aging":  "IT Ops Managers — Proyer, Damon / Cazan, Anca / Gauthier, Guillaume / S K, Arif / Glynn, Colman",
    "catalogue_aging": "IT Ops Managers — Proyer, Damon / Cazan, Anca / Gauthier, Guillaume / S K, Arif / Glynn, Colman",
    "sla_x2":          "IT Ops Managers — Proyer, Damon / Cazan, Anca / Gauthier, Guillaume / S K, Arif / Glynn, Colman",
    "no_movement":     "IT Ops Managers — Proyer, Damon / Cazan, Anca / Gauthier, Guillaume / S K, Arif / Glynn, Colman",
    "repeat_mi":       "IT Ops Managers — per site grouping (identify from Trigger_Log note)",
    "problems_no_rca": "IT Ops Managers — per site grouping",
}

ESCALATION_SUBJECTS = {
    "incident_aging":  "BREACHED — Incident Aging — Recovery Commitment Required",
    "catalogue_aging": "BREACHED — Catalogue Aging — Recovery Commitment Required",
    "sla_x2":          "BREACHED — SLA x2 Violations — Recovery Commitment Required",
    "no_movement":     "WATCH — Stagnant Tickets — Recovery Commitment Required",
    "repeat_mi":       "WATCH — Repeat Major Incidents — Recovery Commitment Required",
    "problems_no_rca": "BREACHED — Problems Without RCA — Recovery Commitment Required",
}


def generate_escalation_body(metric_name: str, status: str, value, note: str, run_date) -> str:
    """Generate a 3-sentence escalation email body from signal data."""
    rule_id = RULE_ID_MAP.get(metric_name, metric_name.upper())
    next_monday = run_date + timedelta(days=(7 - run_date.weekday()))
    next_monday_str = next_monday.strftime("%d %b %Y")
    line1 = f"{rule_id} {status}: {note}."
    line2 = (f"If unresolved by next weekly run ({next_monday_str}), escalation advances "
             f"to EMEA IT Director with formal recovery plan demand.")
    line3 = ("Recovery commitment required: (1) accountable owner (named individual), "
             "(2) committed completion date, (3) throughput commitment, "
             "(4) next checkpoint date.")
    return f"{line1}\n{line2}\n{line3}"


def write_escalation_emails(wb, summary: list, run_date) -> None:
    """
    Clear and rewrite Escalation_Emails sheet — one block per WATCH/BREACHED metric.
    Called inside update_cockpit() before wb.save() to ensure single save.
    """
    if "Escalation_Emails" not in wb.sheetnames:
        logger.warning("Escalation_Emails sheet not found — skipping email generation")
        return

    ws = wb["Escalation_Emails"]

    # Clear all content from row 2 downward (preserve row 1 headers)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    active = [e for e in summary if e["status"] in ("WATCH", "BREACHED")]

    if not active:
        ws.cell(row=2, column=1).value = "Status"
        ws.cell(row=2, column=2).value = (
            f"All metrics GREEN — no escalation required — {run_date.strftime('%d %b %Y')}"
        )
        return

    run_date_str = run_date.strftime("%d %b %Y")
    current_row = 2

    for n, entry in enumerate(active, start=1):
        key     = entry["metric"]
        status  = entry["status"]
        value   = entry["value"]
        note    = entry.get("note", "")
        display = METRIC_DISPLAY_NAMES.get(key, key)
        rule_id = RULE_ID_MAP.get(key, key.upper())
        subject = f"{ESCALATION_SUBJECTS.get(key, f'{status} — {display} — Recovery Commitment Required')} — {run_date_str}"
        body    = generate_escalation_body(key, status, value, note, run_date)

        ws.cell(row=current_row, column=1).value = f"EMAIL {n}"
        ws.cell(row=current_row, column=2).value = f"{display} — {rule_id}"
        current_row += 1

        ws.cell(row=current_row, column=1).value = "To:"
        ws.cell(row=current_row, column=2).value = ESCALATION_RECIPIENTS.get(key, "")
        current_row += 1

        ws.cell(row=current_row, column=1).value = "Subject:"
        ws.cell(row=current_row, column=2).value = subject
        current_row += 1

        ws.cell(row=current_row, column=1).value = "Body:"
        body_cell = ws.cell(row=current_row, column=2)
        body_cell.value = body
        body_cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
        current_row += 1

        current_row += 1  # blank row between blocks

    logger.info(f"Escalation_Emails updated — {len(active)} email(s) written")


def _trigger_log_first_auto_generated_row(ws):
    """
    First 1-based row where column A contains AUTO-GENERATED (manual vs formula block).
    Returns None if the marker is absent.
    """
    for row_idx in range(3, ws.max_row + 2):
        a_val = ws.cell(row=row_idx, column=1).value
        if a_val is not None and "AUTO-GENERATED" in str(a_val).upper():
            return row_idx
    return None


def _trigger_log_has_open_metric(ws, display: str, stop_row: int) -> bool:
    """
    True if any row in [3, stop_row) has column C matching display and H == Open.
    stop_row is exclusive (e.g. first AUTO-GENERATED row).
    """
    want = str(display).strip()
    for row_idx in range(3, stop_row):
        c_val = ws.cell(row=row_idx, column=3).value
        h_val = ws.cell(row=row_idx, column=8).value
        if c_val is None or h_val is None:
            continue
        if str(c_val).strip() != want:
            continue
        if str(h_val).strip().casefold() == "open":
            return True
    return False


def write_trigger_log(wb, results_summary: list):
    """
    Append BREACHED metric entries to the Trigger_Log sheet.
    Skips when an Open row already exists for the same metric (manual section only,
    rows before column A contains AUTO-GENERATED). Also skips same date + same
    metric already logged today.
    Column order: A=Date, B=Domain, C=Metric, D=Trigger Type,
                  E=Threshold, F=Escalated To, G=Action, H=Status, I=Closure Date
    """
    try:
        ws = wb["Trigger_Log"]
    except KeyError:
        logger.warning("Trigger_Log sheet not found in cockpit — skipping trigger log update")
        return

    today_str = datetime.now().strftime("%d/%m/%Y")

    # Find first empty row (scan column A from row 3 onward — row 1=header, row 2=note)
    first_empty = ws.max_row + 1
    for row_idx in range(3, ws.max_row + 2):
        if ws.cell(row=row_idx, column=1).value is None:
            first_empty = row_idx
            break

    auto_row = _trigger_log_first_auto_generated_row(ws)
    open_scan_stop = auto_row if auto_row is not None else first_empty

    # Collect existing entries for today to prevent duplicates
    existing_today = set()
    for row_idx in range(3, first_empty):
        date_val   = ws.cell(row=row_idx, column=1).value
        metric_val = ws.cell(row=row_idx, column=3).value
        if date_val and str(date_val).startswith(today_str[:10]):
            existing_today.add(str(metric_val))

    red_fill   = PatternFill("solid", fgColor="FF0000")
    white_bold = Font(bold=True, color="FFFFFF")

    new_row = first_empty
    logged  = []

    for entry in results_summary:
        if entry.get("status") != "BREACHED":
            continue

        metric  = entry["metric"]
        display = METRIC_DISPLAY_NAMES.get(metric, metric)

        if _trigger_log_has_open_metric(ws, display, open_scan_stop):
            print(f"  Trigger_Log: {display} already Open — skipping duplicate entry")
            continue

        if display in existing_today:
            logger.info(f"Trigger_Log: duplicate skipped for {display} on {today_str}")
            continue

        rule_id   = RULE_ID_MAP.get(metric, "OPS-??")
        threshold = f"{rule_id}: {entry.get('value')} — {entry.get('note', '')}"

        ws.cell(row=new_row, column=1).value = today_str
        ws.cell(row=new_row, column=2).value = "Operations"
        ws.cell(row=new_row, column=3).value = display
        ws.cell(row=new_row, column=4).value = TRIGGER_TYPE_MAP.get(metric, "")
        ws.cell(row=new_row, column=5).value = threshold
        ws.cell(row=new_row, column=6).value = ESCALATION_OWNER_MAP.get(metric, "")
        ws.cell(row=new_row, column=7).value = "Escalation email required — recovery commitment requested"
        status_cell       = ws.cell(row=new_row, column=8)
        status_cell.value = "Open"
        status_cell.fill  = red_fill
        status_cell.font  = white_bold
        ws.cell(row=new_row, column=9).value = None  # Closure Date — blank

        logger.info(f"Trigger_Log: new entry row {new_row} — {display} BREACHED")
        logged.append((today_str, "Operations", display, "Open"))
        existing_today.add(display)
        new_row += 1

    return logged


def update_cockpit(metrics: dict, prev_values: dict, euc_metrics: dict = None):
    """
    Write all calculated metrics into the cockpit Excel file.
    Row references below match the Operations_Panel layout from Build Pack v1.
    Adjust row numbers if your build came out differently.
    """
    try:
        wb = openpyxl.load_workbook(COCKPIT_PATH)
    except FileNotFoundError:
        logger.error(f"Cockpit file not found: {COCKPIT_PATH}")
        print(f"\n  ERROR: Cockpit Excel file not found:")
        print(f"     {COCKPIT_PATH}")
        print(f"     Check COCKPIT_PATH in .env file")
        sys.exit(1)
    except PermissionError:
        logger.warning(f"Permission denied - cockpit file is open: {COCKPIT_PATH}")
        print(f"\n  WARNING: Cannot write to cockpit file - it is currently open")
        print(f"     Close the Excel file and re-run the script")
        print(f"     File: {COCKPIT_PATH}")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Error loading cockpit file: {str(e)}")
        print(f"\n  ERROR: Error loading cockpit Excel file:")
        print(f"     {str(e)}")
        sys.exit(1)

    try:
        ws_ops     = wb["Operations_Panel"]
        ws_physics = wb["Physics_Engine"]
    except KeyError as e:
        logger.error(f"Missing worksheet in cockpit file: {str(e)}")
        print(f"\n  ERROR: Required worksheet not found in cockpit file: {str(e)}")
        print(f"     Expected sheets: Operations_Panel, Physics_Engine")
        sys.exit(1)

    # -- Row mapping: metric -> (Current Value col, Trigger Status col,
    #                           Escalation Required col, Physics Block 4 row)
    # Calibrated from EMEA_Governance_Cockpit_2026.xlsx on 2026-03-07
    ROW_MAP = {
        "incident_aging":   {"cv": "B4",  "ts": "F4",  "er": "G4",  "phys_row": 2, "pct": True},
        "catalogue_aging":  {"cv": "B5",  "ts": "F5",  "er": "G5",  "phys_row": 3, "pct": True},
        "sla_x2":           {"cv": "B6",  "ts": "F6",  "er": "G6",  "phys_row": 4},
        "no_movement":      {"cv": "B8",  "ts": "F8",  "er": "G8",  "phys_row": 5},
        "repeat_mi":        {"cv": "B10", "ts": "F10", "er": "G10", "phys_row": 6},
        "problems_no_rca":  {"cv": "B11", "ts": "F11", "er": "G11", "phys_row": 7},
    }

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    results_summary = []

    for key, m in metrics.items():
        if key not in ROW_MAP:
            continue
        refs = ROW_MAP[key]

        # Current value
        cv_cell = ws_ops[refs["cv"]]
        cv_cell.value = m.get("value")

        # Cell note / comment (openpyxl Comment requires author)
        from openpyxl.comments import Comment
        if m.get("note"):
            cv_cell.comment = Comment(
                f"{now_str}: {m['note']}", "EMEA SDM Refresh"
            )

        # Trigger status
        prev_cv   = prev_values.get(key, {}).get("current")
        prev_4w   = prev_values.get(key, {}).get("avg_4w")
        try:
            prev_cv = float(prev_cv) if prev_cv is not None else None
            prev_4w = float(prev_4w) if prev_4w is not None else None
        except (ValueError, TypeError):
            prev_cv = None
            prev_4w = None
        status    = evaluate_trigger(
            key,
            current    = m.get("value"),
            prev_4w_avg= prev_4w,
            prev_week  = prev_cv,
            extra      = m
        )
        set_trigger_cell(ws_ops, refs["ts"], status)

        # Escalation Required
        esc_cell = ws_ops[refs["er"]]
        esc_cell.value = "YES" if status in ("WATCH", "BREACHED") else "NO"
        esc_cell.fill  = FILLS["BREACHED"] if status in ("WATCH", "BREACHED") \
                         else FILLS["GREEN"]
        esc_cell.font  = FONTS["BREACHED"] if status in ("WATCH", "BREACHED") \
                         else FONTS["GREEN"]

        # Physics Engine Block 4 trend shift
        # Percentage metrics are stored as decimals in Excel (cell format ×100 on display)
        phys_value = m.get("value")
        if refs.get("pct") and phys_value is not None:
            phys_value = round(phys_value / 100, 4)
        shift_physics_trends(ws_physics, refs["phys_row"], phys_value)

        results_summary.append({
            "metric": key,
            "value":  m.get("value"),
            "status": status,
            "note":   m.get("note", "")
        })

    # Update Physics Block 1 (Enterprise_Panel EUC row 4 is applied in main() after this save)
    if euc_metrics and euc_metrics.get("total_units", 0) > 0:
        update_physics_block1(ws_physics, euc_metrics)

    # Append BREACHED entries to Trigger_Log
    write_trigger_log(wb, results_summary)

    # Rewrite Escalation_Emails (single save covers both)
    write_escalation_emails(wb, results_summary, TODAY.date())

    try:
        wb.save(COCKPIT_PATH)
        logger.info(f"Successfully saved cockpit file: {COCKPIT_PATH}")
    except PermissionError:
        logger.warning(f"Permission denied saving cockpit - file is open: {COCKPIT_PATH}")
        print(f"\n  WARNING: Cannot save cockpit file - it was opened during execution")
        print(f"     Close the Excel file and re-run the script")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Error saving cockpit file: {str(e)}")
        print(f"\n  ERROR: Error saving cockpit Excel file:")
        print(f"     {str(e)}")
        sys.exit(1)

    return results_summary


# ---------------------------------------------
# PREVIOUS VALUES (for trend evaluation)
# ---------------------------------------------

def read_prev_values() -> dict:
    """
    Read current cockpit values before overwriting — used for trend comparison.
    Returns dict of {metric_key: {current, avg_4w}}
    Adjust cell refs to match your workbook.
    """
    try:
        wb = openpyxl.load_workbook(COCKPIT_PATH, data_only=True)
        ws = wb["Operations_Panel"]
        return {
    "incident_aging":  {"current": ws["B4"].value,  "avg_4w": ws["C4"].value},
    "catalogue_aging": {"current": ws["B5"].value,  "avg_4w": ws["C5"].value},
    "sla_x2":          {"current": ws["B6"].value,  "avg_4w": ws["C6"].value},
    "no_movement":     {"current": ws["B8"].value,  "avg_4w": ws["C8"].value},
    "repeat_mi":       {"current": ws["B10"].value, "avg_4w": ws["C10"].value},
    "problems_no_rca": {"current": ws["B11"].value, "avg_4w": ws["C11"].value},
}
    except Exception:
        return {}


# ---------------------------------------------
# SHADOW BACKFILL (one-time historical Physics Block 4 Wk2/Wk3)
# ---------------------------------------------

SHADOW_BACKFILL_DONE_PATH = os.path.join(_SCRIPT_DIR, "shadow_backfill.done")
# Plan: 13 Mar / 20 Mar 2026 17:00 — encoded as UTC for SNOW queries (see SNOW_SHADOW_BACKFILL_PLAN.md)
SHADOW_SNAPSHOT_WK2 = datetime(2026, 3, 13, 17, 0, tzinfo=timezone.utc)
SHADOW_SNAPSHOT_WK3 = datetime(2026, 3, 20, 17, 0, tzinfo=timezone.utc)


def _physics_cell_needs_backfill(val) -> bool:
    """True if cell is blank or numeric zero (ghost placeholder)."""
    if val is None:
        return True
    if isinstance(val, str) and not str(val).strip():
        return True
    try:
        n = float(val)
        if abs(n) < 1e-12:
            return True
    except (TypeError, ValueError):
        return False
    return False


def _physics_block4_stored_value(metric_key: str, raw) -> float | int:
    """Match update_cockpit / shift_physics_trends: % as decimals, counts as ints."""
    if metric_key in ("incident_aging", "catalogue_aging"):
        if raw is None:
            return 0.0
        return round(float(raw) / 100.0, 4)
    if raw is None:
        return 0
    return int(raw)


def _compute_metrics_for_as_of(
    incidents: pd.DataFrame,
    mi_history: pd.DataFrame,
    tasks: pd.DataFrame,
    problems: pd.DataFrame,
    as_of: datetime,
) -> dict:
    return {
        "incident_aging": calc_m1_incident_aging(incidents, as_of),
        "catalogue_aging": calc_m2_catalogue_aging(tasks, as_of),
        "sla_x2": calc_m3_sla_x2(incidents, as_of),
        "no_movement": calc_m4_no_movement(incidents, tasks, as_of),
        "repeat_mi": calc_m5_repeat_mi(mi_history),
        "problems_no_rca": calc_m6_problems_no_rca(problems, as_of),
    }


def run_shadow_backfill() -> None:
    """
    One-time backfill: two historical snapshots -> Physics_Engine col C (partial Wk2)
    and col D (full Wk3). Exits if shadow_backfill.done exists next to this script.
    """
    if os.path.isfile(SHADOW_BACKFILL_DONE_PATH):
        print(f"\n  Shadow backfill already completed — remove {SHADOW_BACKFILL_DONE_PATH} to re-run.\n")
        logger.info(f"Shadow backfill skipped — done file exists: {SHADOW_BACKFILL_DONE_PATH}")
        return

    print(f"\n{'='*60}")
    print("SHADOW BACKFILL — historical Physics Block 4 (Wk2 partial / Wk3 full)")
    print(f"{'='*60}\n")
    logger.info("Shadow backfill started")

    global EMEA_SITES, EMEA_LOCATION_FILTER
    print("Loading EMEA site list...")
    sites = fetch_emea_sites()
    EMEA_SITES = sites["site_names"]
    EMEA_LOCATION_FILTER = "location.nameIN" + sites["location_ids"]
    print(f"  Active EMEA sites: {sites['count']} (source: {sites['source']})")

    print(f"\nSnapshot Wk2 (column C partial): {SHADOW_SNAPSHOT_WK2.isoformat()}")
    sig2 = fetch_historical_signal(SHADOW_SNAPSHOT_WK2)
    metrics_wk2 = _compute_metrics_for_as_of(
        sig2["incidents"],
        sig2["mi_history"],
        sig2["tasks"],
        sig2["problems"],
        SHADOW_SNAPSHOT_WK2,
    )

    print(f"\nSnapshot Wk3 (column D full): {SHADOW_SNAPSHOT_WK3.isoformat()}")
    sig3 = fetch_historical_signal(SHADOW_SNAPSHOT_WK3)
    metrics_wk3 = _compute_metrics_for_as_of(
        sig3["incidents"],
        sig3["mi_history"],
        sig3["tasks"],
        sig3["problems"],
        SHADOW_SNAPSHOT_WK3,
    )

    ghost: dict[str, object] = {}
    try:
        wb_ro = openpyxl.load_workbook(COCKPIT_PATH, data_only=True)
        if "Physics_Engine" not in wb_ro.sheetnames:
            print("\n  ERROR: Physics_Engine sheet not found — aborting shadow backfill")
            logger.error("Shadow backfill aborted — Physics_Engine missing")
            return
        ws_ro = wb_ro["Physics_Engine"]
        ghost = {"C3": ws_ro["C3"].value, "C5": ws_ro["C5"].value, "C6": ws_ro["C6"].value}
    except FileNotFoundError:
        print(f"\n  ERROR: Cockpit not found: {COCKPIT_PATH}")
        logger.error("Shadow backfill aborted — cockpit missing")
        return
    except Exception as e:
        print(f"\n  ERROR: Could not read cockpit for ghost check: {e}")
        logger.error(f"Shadow backfill ghost read failed: {e}")
        return

    wk2_writes = [
        ("catalogue_aging", "C3", metrics_wk2["catalogue_aging"].get("value")),
        ("no_movement", "C5", metrics_wk2["no_movement"].get("value")),
        ("repeat_mi", "C6", metrics_wk2["repeat_mi"].get("value")),
    ]

    try:
        wb = openpyxl.load_workbook(COCKPIT_PATH)
    except PermissionError:
        print(f"\n  ERROR: Cannot open cockpit (file may be open): {COCKPIT_PATH}")
        logger.error("Shadow backfill — permission denied loading cockpit")
        return
    except FileNotFoundError:
        print(f"\n  ERROR: Cockpit not found: {COCKPIT_PATH}")
        return
    except Exception as e:
        print(f"\n  ERROR: Loading cockpit failed: {e}")
        logger.error(f"Shadow backfill load error: {e}")
        return

    if "Physics_Engine" not in wb.sheetnames:
        print("\n  ERROR: Physics_Engine sheet not found")
        return

    ws = wb["Physics_Engine"]

    print("\n  Wk2 partial writes (C3/C5/C6) — skip if cell already non-zero:")
    for key, cell, raw in wk2_writes:
        prev = ghost.get(cell)
        if not _physics_cell_needs_backfill(prev):
            print(f"    {cell} ({key}): skip — existing value")
            continue
        stored = _physics_block4_stored_value(key, raw)
        ws[cell] = stored
        print(f"    {cell} ({key}): {stored}")

    wk3_order = [
        "incident_aging",
        "catalogue_aging",
        "sla_x2",
        "no_movement",
        "repeat_mi",
        "problems_no_rca",
    ]
    print("\n  Wk3 full writes (D2–D7):")
    for i, key in enumerate(wk3_order):
        cell = f"D{2 + i}"
        raw = metrics_wk3[key].get("value")
        stored = _physics_block4_stored_value(key, raw)
        ws[cell] = stored
        print(f"    {cell} ({key}): {stored}")

    try:
        wb.save(COCKPIT_PATH)
        with open(SHADOW_BACKFILL_DONE_PATH, "w", encoding="utf-8") as dfh:
            dfh.write(f"completed {datetime.now(timezone.utc).isoformat()}\n")
        logger.info(f"Shadow backfill saved — {COCKPIT_PATH}")
        print(f"\n  Shadow backfill complete — saved {COCKPIT_PATH}")
        print(f"  Done file: {SHADOW_BACKFILL_DONE_PATH}\n")
    except PermissionError:
        logger.error("Shadow backfill save failed — permission denied")
        print("\n  ERROR: Cannot save cockpit — close Excel and re-run")
        return
    except Exception as e:
        logger.error(f"Shadow backfill save failed: {e}")
        print(f"\n  ERROR: Save failed: {e}")
        return


# ---------------------------------------------
# SITE COVERAGE VALIDATION
# ---------------------------------------------

def validate_site_coverage(incidents: pd.DataFrame,
                            tasks: pd.DataFrame,
                            problems: pd.DataFrame):
    """Warn if any of the 25 expected sites are missing from any dataset."""
    all_sites = set()
    for df in [incidents, tasks, problems]:
        if not df.empty and "location" in df.columns:
            all_sites.update(df["location"].dropna().unique())

    missing  = set(EMEA_SITES) - all_sites
    extra    = all_sites - set(EMEA_SITES)

    if missing:
        print(f"  WARNING:  MISSING from data: {sorted(missing)}")
        print("     Check SNOW site filter — these sites returned no records")
    if extra:
        print(f"  WARNING:  UNEXPECTED sites in data: {sorted(extra)}")
        print("     SNOW filter may be too broad")
    if not missing and not extra:
        print("  OK:  All 25 EMEA sites present in data")


# ---------------------------------------------
# MAIN
# ---------------------------------------------

def main():
    # Parse command line arguments
    parser = argparse.ArgumentParser(
        description="EMEA Governance Cockpit — SNOW REST API Refresh Script"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Test mode: fetch and calculate metrics without updating the cockpit Excel file"
    )
    parser.add_argument(
        "--test-fetch",
        action="store_true",
        help="Test incident fetch strategies (caller exclude, location filters, batch-by-site)"
    )
    parser.add_argument(
        "--test-limit",
        type=int,
        default=1000,
        metavar="N",
        help="Record limit per query for --test-fetch (default 1000, use 50000 for full)"
    )
    parser.add_argument(
        "--csv",
        action="store_true",
        help="Load data from CSV instead of SNOW REST API"
    )
    parser.add_argument(
        "--csv-path",
        type=str,
        default=None,
        metavar="PATH",
        help="Full path to consolidated CSV file (default: <script_dir>/SNOW_Exports/Current/EMEA_GOV_Weekly_Consolidated.csv)"
    )
    parser.add_argument(
        "--shadow-backfill",
        action="store_true",
        help="One-time historical backfill (Physics_Engine Block 4 col C partial / D full); creates shadow_backfill.done when complete",
    )
    args = parser.parse_args()

    if args.shadow_backfill:
        run_shadow_backfill()
        return

    if args.test_fetch:
        test_incident_fetch(limit=args.test_limit)
        return

    print(f"\n{'='*60}")
    print(f"EMEA Governance Cockpit — Weekly Refresh")
    if args.dry_run:
        print(f"MODE: DRY RUN (cockpit will NOT be updated)")
    print(f"Run: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'='*60}\n")

    logger.info(f"{'='*60}")
    logger.info(f"EMEA Governance Cockpit refresh started - Mode: {'DRY RUN' if args.dry_run else 'LIVE'}")
    logger.info(f"SNOW Instance: {SNOW_INSTANCE}")

    # 0. Build EMEA site list dynamically from cmn_location (API primary, CSV fallback)
    print("Loading EMEA site list...")
    global EMEA_SITES, EMEA_LOCATION_FILTER
    sites = fetch_emea_sites()
    EMEA_SITES           = sites["site_names"]
    EMEA_LOCATION_FILTER = "location.nameIN" + sites["location_ids"]
    print(f"  Active EMEA sites: {sites['count']} (source: {sites['source']})")
    logger.info(f"EMEA sites: {sites['count']} physical sites | source: {sites['source']}")

    # 1. Read previous cockpit values for trend comparison
    print("Reading previous cockpit values...")
    prev_values = read_prev_values()

    # 2. Fetch data from SNOW or CSV
    if args.csv:
        print("CSV MODE — SNOW REST API bypassed")
        csv_path = args.csv_path or os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "SNOW_Exports", "Current", "EMEA_GOV_Weekly_Consolidated.csv"
        )
        if not os.path.exists(csv_path):
            print(f"ERROR: CSV file not found: {csv_path}")
            print("Export EMEA_GOV_Weekly_Consolidated.csv from SNOW and place it at the path above.")
            return
        incidents, mi_history, tasks, problems = load_from_csv(csv_path)
        logger.info(f"Data source: CSV ({csv_path}) | Run: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    else:
        print("API MODE — querying SNOW REST API...")
        incidents  = fetch_incidents()
        mi_history = fetch_major_incident_history()
        tasks      = fetch_catalogue_tasks()
        problems   = fetch_problems()

    print(f"\n  Incidents (open):    {len(incidents):>5} records")
    print(f"  MI history (60d):    {len(mi_history):>5} records")
    print(f"  Catalogue tasks:     {len(tasks):>5} records")
    print(f"  Problems:            {len(problems):>5} records")

    logger.info(f"Data fetch complete - incidents: {len(incidents)}, MI history: {len(mi_history)}, "
                f"tasks: {len(tasks)}, problems: {len(problems)}")

    # 2b. Fetch EUC/EOSL device replacement data
    print("\nReading EUC/EOSL device replacement data...")
    euc_df = fetch_euc_assets(EUC_PATH)
    euc_metrics = calc_physics_block1_euc(euc_df) if not euc_df.empty else None

    # 3. Validate site coverage
    print("\nValidating site coverage...")
    validate_site_coverage(incidents, tasks, problems)

    # 4. Calculate metrics
    print("\nCalculating metrics...")
    metrics = {
        "incident_aging":  calc_m1_incident_aging(incidents),
        "catalogue_aging": calc_m2_catalogue_aging(tasks),
        "sla_x2":          calc_m3_sla_x2(incidents),
        "no_movement":     calc_m4_no_movement(incidents, tasks),
        "repeat_mi":       calc_m5_repeat_mi(mi_history),
        "problems_no_rca": calc_m6_problems_no_rca(problems),
    }

    # 5. Update cockpit (skip in dry-run mode)
    if args.dry_run:
        # Generate summary without writing to Excel
        summary = []
        for key, m in metrics.items():
            prev_cv   = prev_values.get(key, {}).get("current")
            prev_4w   = prev_values.get(key, {}).get("avg_4w")
            try:
                prev_cv = float(prev_cv) if prev_cv is not None else None
                prev_4w = float(prev_4w) if prev_4w is not None else None
            except (ValueError, TypeError):
                prev_cv = None
                prev_4w = None
            status    = evaluate_trigger(
                key,
                current    = m.get("value"),
                prev_4w_avg= prev_4w,
                prev_week  = prev_cv,
                extra      = m
            )
            summary.append({
                "metric": key,
                "value":  m.get("value"),
                "status": status,
                "note":   m.get("note", "")
            })

        # Print Physics Block 1 EUC/EOSL metrics in dry-run
        if euc_metrics:
            print(f"\n{'-'*60}")
            print("PHYSICS BLOCK 1 — EUC/EOSL DEVICE REPLACEMENT")
            print(f"{'-'*60}")
            print(f"  Total EOSL Units in Scope:       {euc_metrics['total_units']}")
            print(f"  Units Replaced to Date:          (manually maintained)")
            print(f"  Units Remaining:                 {euc_metrics['remaining']}")
            print(f"  Weeks Remaining to 30 Nov 2026:  {euc_metrics['weeks_remaining']}")
            print(f"  Required Weekly Burn:            {euc_metrics['required_weekly_burn']}")
            print(f"  Q4 Adjusted Burn (from 1 Oct):   {euc_metrics['q4_adjusted_burn']}")
            print(f"\n  Note: {euc_metrics['note']}")

        # Dry-run: print Trigger_Log entries that would be written
        today_str = datetime.now().strftime("%d/%m/%Y")
        breached  = [e for e in summary if e.get("status") == "BREACHED"]
        if breached:
            print(f"\n{'-'*60}")
            print("TRIGGER_LOG ENTRIES (dry-run — not written)")
            print(f"{'-'*60}")
            print(f"  {'Date':<12} {'Domain':<12} {'Metric':<25} {'Status'}")
            print(f"  {'-'*10} {'-'*10} {'-'*23} {'-'*8}")
            for e in breached:
                display = METRIC_DISPLAY_NAMES.get(e["metric"], e["metric"])
                print(f"  {today_str:<12} {'Operations':<12} {display:<25} BREACHED")

    else:
        print(f"\nUpdating cockpit: {COCKPIT_PATH}")
        summary = update_cockpit(metrics, prev_values, euc_metrics)
        save_enterprise_panel_euc_after_cockpit(COCKPIT_PATH, euc_metrics)

    # 6. Print summary
    print(f"\n{'-'*60}")
    print("REFRESH SUMMARY")
    print(f"{'-'*60}")
    breached = [r for r in summary if r["status"] == "BREACHED"]
    watch    = [r for r in summary if r["status"] == "WATCH"]

    for r in summary:
        icon = "[BREACHED]" if r["status"] == "BREACHED" \
               else "[WATCH]" if r["status"] == "WATCH" \
               else "[GREEN]"
        print(f"  {icon} {r['metric']:<22} {str(r['value']):<10} {r['status']}")
        if r["note"]:
            print(f"       {r['note']}")

    print(f"\n  BREACHED: {len(breached)} | WATCH: {len(watch)} "
          f"| GREEN: {len(summary) - len(breached) - len(watch)}")

    if breached:
        print("\n  WARNING:  BREACHED metrics require immediate escalation:")
        for r in breached:
            print(f"     -> {r['metric']}: {r['note']}")

    # Log final summary
    green_count = len(summary) - len(breached) - len(watch)
    logger.info(f"Refresh complete - Metrics updated: {len(summary)} | "
                f"BREACHED: {len(breached)} | WATCH: {len(watch)} | GREEN: {green_count}")

    if breached:
        logger.warning(f"BREACHED metrics requiring escalation: {', '.join([r['metric'] for r in breached])}")

    print(f"\n{'='*60}")
    if args.dry_run:
        print("DRY RUN — cockpit not updated")
        logger.info("Dry run completed successfully - no cockpit update")
    else:
        print("Cockpit saved. Refresh complete.")
        logger.info("Live run completed successfully - cockpit updated")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
