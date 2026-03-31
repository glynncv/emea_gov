"""
One-time population of Physics_Engine Block 3 (OT Remediation Stagnation Tracker), rows 28–39.

Spec: CURSOR_PROMPT_Physics_Block3_OT.md

Usage:
    python scripts/populate_physics_block3_ot.py
    python scripts/populate_physics_block3_ot.py --cockpit "C:\\path\\to\\EMEA_Governance_Cockpit_2026.xlsx"
    python scripts/populate_physics_block3_ot.py --dry-run

Requires: openpyxl, python-dotenv (same as emea_gov_refresh.py)
"""

from __future__ import annotations

import argparse
import os
import shutil
import sys
import textwrap
from datetime import datetime
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

load_dotenv()

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_REPO_ROOT = os.path.dirname(_SCRIPT_DIR)
COCKPIT_DEFAULT = os.getenv(
    "COCKPIT_PATH", os.path.join(_REPO_ROOT, "EMEA_Governance_Cockpit_2026.xlsx")
)

SHEET = "Physics_Engine"

# Rows 30–35: active EMEA 2026 OT sites (col A name, col D funding)
ACTIVE_SITES: list[tuple[int, str, str]] = [
    (30, "Amal - Sweden", "TBC"),
    (31, "Gillingham - United Kingdom", "TBC"),
    (32, "Iasi - Romania", "TBC"),
    (33, "Izmir - Turkey / ESBAS 2 (PT Phase 1)", "TBC"),
    (34, "Izmir - Turkey / ESBAS 3 (PT Phase 2)", "TBC"),
    (35, "Technical Center Krakow", "TBC"),
]

# Rows 36–38: out-of-scope summary (col A label; col D N/A; F/G literals)
OUT_OF_SCOPE: list[tuple[int, str]] = [
    (
        36,
        "[OUT OF SCOPE 2026] Blois, Hartridge, Warwick, Stonehouse, Krakow Office — 2027",
    ),
    (
        37,
        "[OUT OF SCOPE 2026] Bucharest, Wroclaw, Cergy, Rzeszow, Warwick Titan, Madrid, Duesseldorf, Belval, Dubai — 2028",
    ),
    (
        38,
        "[OUT OF SCOPE] Cinisello, Istanbul, Warsaw (Delphi Academy), Warsaw — no OT devices confirmed",
    ),
]

YELLOW = PatternFill("solid", fgColor="FFFF00")
WHITE = PatternFill("solid", fgColor="FFFFFF")
GREY = PatternFill("solid", fgColor="D9D9D9")

FONT_ITALIC = Font(italic=True)


def _formula_days_since(row: int) -> str:
    return f'=IF(B{row}="","",TODAY()-B{row})'


def _formula_stagnation(row: int) -> str:
    return (
        f'=IF(C{row}="","",IF(C{row}>30,"STAGNANT",IF(C{row}>14,"AT RISK","OK")))'
    )


def _formula_status(row: int) -> str:
    return (
        f'=IF(C{row}="","PENDING",IF(OR(E{row}>=2,C{row}>30),"BREACHED",'
        f'IF(C{row}>14,"WATCH","ON TRAJECTORY")))'
    )


def _remove_cf_for_range(ws: Worksheet, range_str: str) -> None:
    try:
        del ws.conditional_formatting[range_str]
    except KeyError:
        pass


def _apply_conditional_formatting_g(ws: Worksheet) -> None:
    """Col G rows 30–35: status-based fills (overrides base white)."""
    rng = "G30:G35"
    _remove_cf_for_range(ws, rng)

    # Top-left cell for relative CF references (Excel adjusts per row)
    ref = "G30"

    rules: list[tuple[str, PatternFill, Font]] = [
        (
            "BREACHED",
            PatternFill("solid", fgColor="FF0000"),
            Font(bold=True, color="FFFFFF"),
        ),
        (
            "WATCH",
            PatternFill("solid", fgColor="FFC000"),
            Font(bold=True, color="000000"),
        ),
        (
            "ON TRAJECTORY",
            PatternFill("solid", fgColor="00B050"),
            Font(bold=True, color="FFFFFF"),
        ),
        (
            "PENDING",
            PatternFill("solid", fgColor="D9D9D9"),
            Font(color="000000"),
        ),
    ]

    for text, fill, font in rules:
        # Escape double quotes inside text for Excel formula
        safe = text.replace('"', '""')
        formula = [f'{ref}="{safe}"']
        ws.conditional_formatting.add(
            rng, FormulaRule(formula=formula, fill=fill, font=font)
        )


def populate_block3(ws: Worksheet) -> None:
    """Write Block 3 data rows 30–39 (headers 28–29 untouched)."""

    # --- Active rows 30–35 ---
    for row, site_name, funding in ACTIVE_SITES:
        c = ws.cell(row=row, column=1)
        c.value = site_name
        c.fill = WHITE
        c.font = Font()
        c.alignment = Alignment()

        b = ws.cell(row=row, column=2)
        b.value = None
        b.fill = YELLOW
        b.font = Font()

        ws.cell(row=row, column=3).value = _formula_days_since(row)
        ws.cell(row=row, column=3).fill = WHITE
        ws.cell(row=row, column=3).font = Font()

        d = ws.cell(row=row, column=4)
        d.value = funding
        d.fill = YELLOW
        d.font = Font()

        e = ws.cell(row=row, column=5)
        e.value = None
        e.fill = YELLOW
        e.font = Font()

        ws.cell(row=row, column=6).value = _formula_stagnation(row)
        ws.cell(row=row, column=6).fill = WHITE
        ws.cell(row=row, column=6).font = Font()

        ws.cell(row=row, column=7).value = _formula_status(row)
        ws.cell(row=row, column=7).fill = WHITE
        ws.cell(row=row, column=7).font = Font()

    # --- Out-of-scope rows 36–38 ---
    for row, label in OUT_OF_SCOPE:
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = GREY
            cell.font = FONT_ITALIC
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = None
        ws.cell(row=row, column=3).value = None
        ws.cell(row=row, column=4).value = "N/A"
        ws.cell(row=row, column=5).value = None
        ws.cell(row=row, column=6).value = "N/A"
        ws.cell(row=row, column=7).value = "OUT OF SCOPE"

    # --- Spare row 39 ---
    for col in range(1, 8):
        cell = ws.cell(row=39, column=col)
        cell.value = None
        cell.fill = WHITE
        cell.font = Font()
        cell.alignment = Alignment()

    _apply_conditional_formatting_g(ws)


def _normalize_cockpit_path(raw: str) -> str:
    """Strip whitespace / accidental line breaks from .env values."""
    s = "".join(raw.strip().splitlines())
    return os.path.normpath(s)


def _wrap_width() -> int:
    """Target line width for wrapped messages (narrow terminals)."""
    try:
        return max(52, min(100, os.get_terminal_size().columns - 2))
    except OSError:
        return 80


def _print_wrapped(prefix: str, body: str) -> None:
    """Print prefix + body with explicit wrapping (avoids mid-path console wrap)."""
    width = _wrap_width()
    line = f"{prefix}{body}"
    if len(line) <= width:
        print(line, flush=True)
        return
    wrapped = textwrap.fill(
        body,
        width=width,
        initial_indent=prefix,
        subsequent_indent=" " * len(prefix),
        break_long_words=True,
        break_on_hyphens=False,
    )
    print(wrapped, flush=True)


def run(
    cockpit_path: str, dry_run: bool = False, no_backup: bool = False
) -> int:
    cockpit_path = _normalize_cockpit_path(cockpit_path)
    if not os.path.isfile(cockpit_path):
        print(f"ERROR: Cockpit file not found: {cockpit_path}", file=sys.stderr)
        return 1

    if dry_run:
        cp = Path(cockpit_path)
        print("DRY RUN — cockpit file:")
        _print_wrapped("  Folder: ", str(cp.parent))
        print(f"  File:   {cp.name}", flush=True)
        wb = openpyxl.load_workbook(cockpit_path)
    else:
        if not no_backup:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            base, ext = os.path.splitext(cockpit_path)
            backup_path = f"{base}_BACKUP_Block3_{ts}{ext}"
            shutil.copy2(cockpit_path, backup_path)
            _print_wrapped("Backup written: ", backup_path)
        wb = openpyxl.load_workbook(cockpit_path)

    if SHEET not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET}' not found.", file=sys.stderr)
        return 1

    ws = wb[SHEET]
    populate_block3(ws)

    if dry_run:
        print(
            "DRY RUN — Block 3 updates applied in memory only; workbook not saved.",
            flush=True,
        )
        return 0

    wb.save(cockpit_path)
    _print_wrapped("Saved: ", cockpit_path)
    return 0


def main() -> int:
    p = argparse.ArgumentParser(description="Populate Physics_Engine Block 3 (OT Remediation).")
    p.add_argument(
        "--cockpit",
        default=COCKPIT_DEFAULT,
        help=f"Cockpit .xlsx path (default: COCKPIT_PATH or {COCKPIT_DEFAULT})",
    )
    p.add_argument(
        "--dry-run",
        action="store_true",
        help="Load workbook and apply in memory only; do not save or write backup.",
    )
    p.add_argument(
        "--no-backup",
        action="store_true",
        help="Skip timestamped backup copy before save.",
    )
    args = p.parse_args()
    return run(args.cockpit, dry_run=args.dry_run, no_backup=args.no_backup)


if __name__ == "__main__":
    raise SystemExit(main())
